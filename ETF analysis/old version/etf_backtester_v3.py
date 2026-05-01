"""
etf_backtester_v3.0 - 專業實戰版
1. 導入 Walk-forward Validation (WFV) 滾動驗證
2. 導入 Optuna 自動參數優化
3. 嚴格保留原始 Excel 結構與視覺格式
"""
import pandas as pd
import numpy as np
import optuna
from xgboost import XGBRegressor
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
import os, json
from datetime import datetime

# 抑制 Optuna 輸出日誌，保持終端機乾淨
optuna.logging.set_verbosity(optuna.logging.WARNING)

# ==========================================
# ETF 資訊 (保持不變)
# ==========================================
ETF_NAMES = {
    '0050':'元大台灣50', '00713':'元大台灣高息低波', '00762':'元大全球AI', 
    '00965':'元大航太防衛科技', '009816':'凱基台灣TOP50', '00878':'國泰永續高股息',
    '00929':'復華台灣科技優息', '0056':'元大高股息', '00919':'群益台灣精選高息', 
    '00881':'國泰台灣科技龍頭'
}

ETF_CATEGORY = {
    '0050':'市值型', '009816':'市值型', '00881':'科技型', '00762':'主題型', 
    '00965':'主題型', '00713':'高息低波', '0056':'高息型', '00878':'高息ESG', 
    '00919':'高息型', '00929':'科技高息'
}

TECH_FEATURES = ['收盤價','成交量','MA5','MA20','ret5','RSI14','ATR','BB_width','BIAS20','量能比','MACD_hist']
MACRO_FEATURES = ['台幣匯率','VIX指數','RSP','SP500_ret','DXY_ret','美債10Y_ret','黃金_ret','原油_ret']
TYPE_FEATURES  = ['ETFType','rolling_beta_60d','beta_regime','market_stress','beta_sp500_static']
VAL_FEATURES   = ['price_zscore','price_trend']

ALL_FEATURES = list(dict.fromkeys(TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES + VAL_FEATURES))

# ==========================================
# WFV + Optuna 核心邏輯[cite: 2]
# ==========================================
def build_features(df, code):
    d = df.copy()
    d['日期'] = pd.to_datetime(d['日期'])
    d = d.sort_values('日期').reset_index(drop=True)
    close = d['收盤價'].replace(0, np.nan)
    d['MA5'] = close.rolling(5, min_periods=1).mean()
    d['MA20'] = close.rolling(20, min_periods=1).mean()
    d['ret5'] = close.pct_change(5)
    d['BIAS20'] = (close - d['MA20']) / d['MA20']
    for col in ALL_FEATURES:
        if col not in d.columns: d[col] = 0.0
    return d.ffill().bfill()

def train_optuna_model(Xs, ys):
    if len(ys) < 20: return None
    
    def objective(trial):
        param = {
            'n_estimators': trial.suggest_int('n_estimators', 50, 120),
            'max_depth': trial.suggest_int('max_depth', 2, 4),
            'learning_rate': trial.suggest_float('learning_rate', 0.01, 0.15),
            'subsample': trial.suggest_float('subsample', 0.7, 0.9),
            'random_state': 42,
            'n_jobs': -1,
            'verbosity': 0
        }
        # 簡單切分驗證集
        split = int(len(Xs) * 0.8)
        X_t, X_v = Xs[:split], Xs[split:]
        y_t, y_v = ys[:split], ys[split:]
        
        reg = XGBRegressor(**param)
        reg.fit(X_t, y_t)
        return np.mean((reg.predict(X_v) - y_v)**2)

    study = optuna.create_study(direction='minimize')
    study.optimize(objective, n_trials=10) # 為了速度設定 10 次
    
    best_mdl = Pipeline([
        ('s', StandardScaler()), 
        ('m', XGBRegressor(**study.best_params, random_state=42))
    ])
    best_mdl.fit(Xs, ys)
    return best_mdl

# ==========================================
# 視覺呈現工具 (完全保留自 v1/v2)[cite: 2]
# ==========================================
C_NAVY, C_BLUE, C_WHITE = '1F3864', '2E75B6', 'FFFFFF'
def _thin(): return Border(left=Side(style='thin',color='BFBFBF'),right=Side(style='thin',color='BFBFBF'),top=Side(style='thin',color='BFBFBF'),bottom=Side(style='thin',color='BFBFBF'))
def _hdr(cell, val, bg=C_BLUE, size=10):
    cell.value, cell.font, cell.fill = val, Font(bold=True, color=C_WHITE, size=size), PatternFill('solid', fgColor=bg)
    cell.alignment, cell.border = Alignment(horizontal='center',vertical='center',wrap_text=True), _thin()
def tier_style(wr):
    if wr >= 0.65:   return PatternFill('solid',fgColor='C6EFCE'), Font(color='006100',bold=True), '強'
    elif wr >= 0.55: return PatternFill('solid',fgColor='DDEBF7'), Font(color='1F497D',bold=True), '中'
    else:            return PatternFill('solid',fgColor='FFC7CE'), Font(color='9C0006',bold=True), '弱'
def auto_width(ws):
    for col in ws.columns:
        best = 12
        for cell in col:
            if cell.value: best = min(60, max(best, len(str(cell.value))+4))
        ws.column_dimensions[get_column_letter(col[0].column)].width = best

# ==========================================
# 主執行程序
# ==========================================
HORIZONS = {'3M':63, '6M':126, '12M':252}
DB_FILE, OUTPUT = 'ETF歷史分析資料庫.xlsx', 'ETF回測報告_v2.xlsx'

def main():
    if not os.path.exists(DB_FILE): print("找不到資料庫"); return
    xl = pd.ExcelFile(DB_FILE)
    all_dfs = {s.split()[0].strip(): build_features(pd.read_excel(DB_FILE, sheet_name=s), s) for s in xl.sheet_names}
    
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = '回測準確率總覽'
    _hdr(ws_sum['A1'], f'ETF AI 實戰回測報告 (WFV+Optuna)　{datetime.now().strftime("%Y-%m-%d")}', bg=C_NAVY, size=13)
    ws_sum.merge_cells('A1:L1')
    
    hdrs = ['代號與名稱','類型','預測力','12M勝率','3M完全(±10%)','3M方向勝率','6M完全(±10%)','6M方向勝率','12M完全(±10%)','12M方向勝率','樣本數','特徵數']
    for ci, h in enumerate(hdrs, 1): _hdr(ws_sum.cell(row=2, column=ci), h)

    global_stats, results_summary = {hz:{'perf':[],'trend':[],'n':[]} for hz in HORIZONS}, []

    for code, df_feat in all_dfs.items():
        print(f"正在分析 {code}...")
        display = f"{code} {ETF_NAMES.get(code, code)}"
        ws_det = wb.create_sheet(display[:31])
        _hdr(ws_det['A1'], f'{display}　WFV 滾動回測明細', bg=C_NAVY, size=12)
        ws_det.merge_cells('A1:H1')
        for ci, h in enumerate(['週期','完全成功(±10%)','方向勝率','樣本數','平均預測漲幅','平均實際漲幅','加權評分'],1): _hdr(ws_det.cell(row=2,column=ci), h)

        rec, det_row = {'name':display, 'code':code, 'acc_info':{}, 'trend_info':{}, 'total_weight':0, 'feat_count':len(ALL_FEATURES)}, 3

        # 聯合訓練集 (排除當前 ETF)
        others = [d for c, d in all_dfs.items() if c != code]

        for hz_label, hz_days in HORIZONS.items():
            # 實施 WFV 滾動[cite: 2]
            # 取資料的最後 30% 作為 WFV 測試期，確保結果真實
            test_start = int(len(df_feat) * 0.7)
            all_preds, all_acts = [], []
            
            # 滾動步長
            for i in range(test_start, len(df_feat) - hz_days, 63):
                # 訓練集：使用其他 ETF 在「當時」之前的資料
                train_pool = [d.iloc[:i] for d in others]
                Xs_list, ys_list = [], []
                for td in train_pool:
                    td['target'] = td['收盤價'].pct_change(hz_days).shift(-hz_days)
                    v = td[ALL_FEATURES + ['target']].dropna()
                    Xs_list.append(v[ALL_FEATURES].values); ys_list.append(v['target'].values)
                
                if not Xs_list: continue
                model = train_optuna_model(np.vstack(Xs_list), np.concatenate(ys_list))
                
                if model:
                    # 預測當前 ETF 在該時間點的表現
                    curr_x = df_feat.iloc[i:i+1][ALL_FEATURES].values
                    curr_y = (df_feat['收盤價'].iloc[i+hz_days] - df_feat['收盤價'].iloc[i]) / df_feat['收盤價'].iloc[i]
                    all_preds.append(model.predict(curr_x)[0])
                    all_acts.append(curr_y)

            if not all_preds: continue
            
            all_preds, all_acts = np.array(all_preds), np.array(all_acts)
            perf = float(np.mean(np.abs(all_acts - all_preds) <= 0.10))
            trend = float(np.mean((all_preds * all_acts) > 0))
            n = len(all_preds)

            rec['acc_info'][hz_label], rec['trend_info'][hz_label] = f"{perf:.1%}", f"{trend:.1%}"
            rec['total_weight'] += n
            global_stats[hz_label]['perf'].append(perf); global_stats[hz_label]['trend'].append(trend); global_stats[hz_label]['n'].append(n)

            for ci, v in enumerate([hz_label, f"{perf:.1%}", f"{trend:.1%}", n, f"{all_preds.mean():.2%}", f"{all_acts.mean():.2%}", f"{(perf*0.4+trend*0.6):.3f}"], 1):
                ws_det.cell(row=det_row, column=ci, value=v)
            det_row += 1
        
        auto_width(ws_det); results_summary.append(rec)

    # 填寫總覽 (視覺格式嚴格繼承)[cite: 2]
    for ri, rec in enumerate(results_summary, 3):
        ws_sum.cell(row=ri, column=1, value=rec['name']).font = Font(bold=True)
        ws_sum.cell(row=ri, column=2, value=ETF_CATEGORY.get(rec['code'],'市值型'))
        wr12 = float(rec['trend_info'].get('12M','0%').rstrip('%'))/100
        stars = '★★★★★' if wr12>=0.75 else ('★★★★☆' if wr12>=0.6 else ('★★★☆☆' if wr12>=0.5 else '★★☆☆☆'))
        fill, font, label = tier_style(wr12)
        c_p = ws_sum.cell(row=ri, column=3, value=f"{stars} {label}")
        c_p.fill, c_p.font, c_p.alignment = fill, font, Alignment(horizontal='center')
        ws_sum.cell(row=ri, column=4, value=f"{wr12:.1%}")
        col = 5
        for hz in ['3M','6M','12M']:
            ws_sum.cell(row=ri, column=col, value=rec['acc_info'].get(hz,'-'))
            cell_t = ws_sum.cell(row=ri, column=col+1, value=rec['trend_info'].get(hz,'-'))
            if '-' not in cell_t.value:
                f, fn, _ = tier_style(float(cell_t.value.rstrip('%'))/100)
                cell_t.fill, cell_t.font = f, fn
            col += 2
        ws_sum.cell(row=ri, column=11, value=rec['total_weight'])
        ws_sum.cell(row=ri, column=12, value=rec['feat_count'])

    # 加權平均列[cite: 2]
    last_row = len(results_summary) + 4
    ws_sum.cell(row=last_row, column=1, value="WFV 實戰加權平均").font = Font(bold=True, color=C_WHITE)
    ws_sum.cell(row=last_row, column=1).fill = PatternFill('solid', fgColor=C_NAVY)
    col = 5
    for hz in ['3M','6M','12M']:
        if global_stats[hz]['n']:
            wp = np.average(global_stats[hz]['perf'], weights=global_stats[hz]['n'])
            wt = np.average(global_stats[hz]['trend'], weights=global_stats[hz]['n'])
            ws_sum.cell(row=last_row, column=col, value=f"{wp:.1%}").font = Font(bold=True, color='FF0000')
            ws_sum.cell(row=last_row, column=col+1, value=f"{wt:.1%}").font = Font(bold=True, color='FF0000')
        col += 2

    auto_width(ws_sum); wb.save(OUTPUT)
    print(f"✨ WFV+Optuna 回測完成！報告：{OUTPUT}")

if __name__ == "__main__":
    main()
