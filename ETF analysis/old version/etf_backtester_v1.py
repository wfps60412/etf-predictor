"""
etf_backtester.py  v1.0
ETF 跨標的聯合訓練回測

特徵設計：ETF 無 PER/月營收/外資籌碼，改用：
  TECH:  技術指標（11個）
  MACRO: 總經指標（6個）
  TYPE:  ETF類型 + Beta（5個）
  VAL:   估值動能（2個，price_zscore/price_trend）

依 ETF 類型使用不同特徵組合：
  市值型/科技型/主題型：TECH + MACRO + TYPE
  高息型/高息低波/高息ESG/科技高息：TECH + MACRO + TYPE + VAL
"""
import pandas as pd
import numpy as np
from xgboost import XGBRegressor
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
import argparse, os, json, ast
from datetime import datetime

# ==========================================
# ETF 資訊
# ==========================================
ETF_NAMES = {
    '0050':'元大台灣50',       '00713':'元大台灣高息低波',
    '00762':'元大全球AI',       '00965':'元大航太防衛科技',
    '009816':'凱基台灣TOP50',   '00878':'國泰永續高股息',
    '00929':'復華台灣科技優息', '0056':'元大高股息',
    '00919':'群益台灣精選高息', '00881':'國泰台灣科技龍頭',
}

ETF_CATEGORY = {
    '0050':'市值型',   '009816':'市值型',  '00881':'科技型',
    '00762':'主題型',  '00965':'主題型',   '00713':'高息低波',
    '0056':'高息型',   '00878':'高息ESG',  '00919':'高息型',
    '00929':'科技高息',
}

# ==========================================
# 特徵群組
# ==========================================
TECH_FEATURES = ['收盤價','成交量','MA5','MA20','ret5','RSI14',
                 'ATR14','BB_width','BIAS20','量能比','MACD_hist']
MACRO_FEATURES = ['台幣匯率','VIX指數','RSP','SP500_ret','DXY_ret']
TYPE_FEATURES  = ['ETFType','rolling_beta_60d','beta_regime',
                  'market_stress','beta_sp500_static']
VAL_FEATURES   = ['price_zscore','price_trend']

# ETF 類型 → 特徵組合
CATEGORY_FEATURES = {
    '市值型':   TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES,
    '科技型':   TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES,
    '主題型':   TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES,
    '高息型':   TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES + VAL_FEATURES,
    '高息低波': TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES + VAL_FEATURES,
    '高息ESG':  TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES + VAL_FEATURES,
    '科技高息': TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES + VAL_FEATURES,
}
DEFAULT_FEATURES = TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES

ALL_FEATURES = list(dict.fromkeys(
    TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES + VAL_FEATURES))

# ==========================================
# 特徵工程
# ==========================================
def build_features(df: pd.DataFrame, code: str) -> pd.DataFrame:
    d = df.copy()
    d['日期'] = pd.to_datetime(d['日期'])
    d = d.sort_values('日期').reset_index(drop=True)

    close = d['收盤價'].replace(0, np.nan)
    vol   = d['成交量'].replace(0, np.nan)

    d['MA5']    = close.rolling(5).mean().round(4)
    d['MA20']   = close.rolling(20).mean().round(4)
    d['ret5']   = close.pct_change(5).round(4)
    d['BIAS20'] = ((close - d['MA20']) / d['MA20'].replace(0,np.nan)).round(4)

    delta = close.diff()
    gain  = delta.clip(lower=0).rolling(14).mean()
    loss  = (-delta.clip(upper=0)).rolling(14).mean()
    d['RSI14'] = (100 - 100/(1+gain/loss.replace(0,np.nan))).round(4)

    high  = d['最高價'].replace(0, np.nan)
    low   = d['最低價'].replace(0, np.nan)
    hl    = high - low
    hc    = (high - close.shift(1)).abs()
    lc    = (low  - close.shift(1)).abs()
    tr    = pd.concat([hl,hc,lc],axis=1).max(axis=1)
    d['ATR14'] = tr.rolling(14).mean().round(4)

    bb_mid = close.rolling(20).mean()
    bb_std = close.rolling(20).std()
    d['BB_width'] = ((bb_mid + 2*bb_std - (bb_mid - 2*bb_std))
                     / bb_mid.replace(0,np.nan)).round(4)

    vol_ma20 = vol.rolling(20).mean()
    d['量能比'] = (vol / vol_ma20.replace(0,np.nan)).round(4)

    ema12 = close.ewm(span=12,adjust=False).mean()
    ema26 = close.ewm(span=26,adjust=False).mean()
    macd  = ema12 - ema26
    signal= macd.ewm(span=9,adjust=False).mean()
    d['MACD_hist'] = (macd - signal).round(4)

    # 補缺失欄位
    for col in ['RSP','SP500_ret','DXY_ret','台幣匯率','VIX指數',
                'ETFType','rolling_beta_60d','beta_regime',
                'market_stress','beta_sp500_static',
                'price_zscore','price_trend']:
        if col not in d.columns:
            d[col] = 0.0

    return d.dropna(subset=['收盤價','MA5','MA20'])

# ==========================================
# 跨標的聯合訓練
# ==========================================
def build_model(other_dfs, hz_days, feat_cols=None):
    if feat_cols is None:
        feat_cols = ALL_FEATURES
    Xs, ys = [], []
    for df in other_dfs:
        tmp = df.copy()
        tmp['target'] = tmp['收盤價'].pct_change(hz_days).shift(-hz_days)
        fc  = [f for f in feat_cols if f in tmp.columns]
        v   = tmp[fc+['target']].dropna()
        v   = v[v['target'].abs() <= 0.6]
        if len(v) > 20:
            Xs.append(v[fc].values)
            ys.append(v['target'].values)
    if not Xs:
        return None, []
    Xa = np.vstack(Xs); ya = np.concatenate(ys)
    m   = Xa.std(axis=0) > 0
    fc2 = [f for f,vv in zip(fc, m) if vv]
    Xa  = Xa[:,m]
    mdl = Pipeline([('s',StandardScaler()),
                    ('m',XGBRegressor(n_estimators=200,max_depth=4,
                                      learning_rate=0.05,subsample=0.8,
                                      colsample_bytree=0.8,random_state=42,
                                      verbosity=0))])
    mdl.fit(Xa, ya)
    return mdl, fc2

# ==========================================
# Excel 格式
# ==========================================
C_NAVY = '1F3864'; C_BLUE = '2E75B6'; C_WHITE = 'FFFFFF'

def _thin():
    s = Side(style='thin',color='BFBFBF')
    return Border(left=s,right=s,top=s,bottom=s)

def _hdr(cell, val, bg=C_BLUE, size=10):
    cell.value = val
    cell.font  = Font(bold=True, color=C_WHITE, size=size)
    cell.fill  = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
    cell.border = _thin()

def _pct(v):
    try: return f"{float(str(v).rstrip('%'))/100:.1%}" if '%' in str(v) else f"{float(v):.1%}"
    except: return str(v)

def tier_style(wr):
    if wr >= 0.65:   return PatternFill('solid',fgColor='C6EFCE'), Font(color='006100',bold=True), 'HIGH ●'
    elif wr >= 0.55: return PatternFill('solid',fgColor='DDEBF7'), Font(color='1F497D',bold=True), 'MID  ●'
    else:            return PatternFill('solid',fgColor='FFC7CE'), Font(color='9C0006',bold=True), 'LOW  ●'

def auto_width(ws, min_w=12, max_w=60):
    for col in ws.columns:
        best = min_w
        for cell in col:
            if cell.value:
                best = min(max_w, max(best, len(str(cell.value))+4))
        ws.column_dimensions[get_column_letter(col[0].column)].width = best

# ==========================================
# 回測主體
# ==========================================
HORIZONS = {'3M':63, '6M':126, '12M':252}
OUTPUT    = 'ETF回測報告_v1.xlsx'
DB_FILE   = 'ETF歷史分析資料庫.xlsx'

def run_backtest():
    print("="*60)
    print("  ETF 回測系統 v1.0")
    print("="*60)

    if not os.path.exists(DB_FILE):
        print(f"❌ 找不到 {DB_FILE}，請先執行 etf_analysis.py")
        return

    xl = pd.ExcelFile(DB_FILE)
    all_feats = {}
    for sheet in xl.sheet_names:
        code = sheet.split()[0].strip()
        try:
            df_raw = pd.read_excel(DB_FILE, sheet_name=sheet)
            all_feats[code] = build_features(df_raw, code)
            print(f"  ✅ {sheet}: {len(all_feats[code])} 筆特徵")
        except Exception as e:
            print(f"  ❌ {sheet}: {e}")

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = '回測準確率總覽'

    # 標題
    n_cols = 15
    ws_sum.merge_cells(f'A1:{get_column_letter(n_cols)}1')
    _hdr(ws_sum['A1'], f'ETF AI 回測報告 v1.0　{datetime.now().strftime("%Y-%m-%d")}', bg=C_NAVY, size=13)
    ws_sum.row_dimensions[1].height = 28

    hdrs = ['代號與名稱','類型','預測力','12M勝率',
            '3M完全(±10%)','3M方向勝率',
            '6M完全(±10%)','6M方向勝率',
            '12M完全(±10%)','12M方向勝率',
            '樣本數','特徵數']
    for ci, h in enumerate(hdrs, 1):
        _hdr(ws_sum.cell(row=2, column=ci), h)
    ws_sum.row_dimensions[2].height = 36

    global_stats = {hz:{'perf':[],'trend':[],'n':[]} for hz in HORIZONS}
    results_summary = []
    power_data = {}

    for code, df_feat in all_feats.items():
        name     = ETF_NAMES.get(code, code)
        category = ETF_CATEGORY.get(code, '市值型')
        feat_cols = CATEGORY_FEATURES.get(category, DEFAULT_FEATURES)
        others   = [d for c,d in all_feats.items() if c != code]
        display  = f"{code} {name}"

        print(f"\n📊 {display}  [{category} / {len(feat_cols)}特徵]")

        ws_det = wb.create_sheet(display[:31])
        ws_det.merge_cells('A1:H1')
        _hdr(ws_det['A1'], f'{display}　回測明細', bg=C_NAVY, size=12)
        ws_det.row_dimensions[1].height = 24
        for ci, h in enumerate(['週期','完全成功(±10%)','方向勝率','樣本數',
                                  '平均預測漲幅','平均實際漲幅','加權評分'],1):
            _hdr(ws_det.cell(row=2,column=ci), h)

        rec = {'name':display,'acc_info':{},'trend_info':{},'final_score':0,
               'total_weight':0,'feat_count':0}
        det_row = 3

        for hz_label, hz_days in HORIZONS.items():
            model, feat_names = build_model(others, hz_days, feat_cols)
            if model is None:
                continue
            rec['feat_count'] = len(feat_names)

            dt = df_feat.copy()
            dt['ar'] = (dt['收盤價'].shift(-hz_days)-dt['收盤價'])/dt['收盤價']
            vt = dt[[f for f in feat_names if f in dt.columns]+['ar']].dropna()
            vt = vt[vt['ar'].abs()<=0.6]
            if len(vt)<10: continue

            te   = vt.iloc[::3]
            fc3  = [f for f in feat_names if f in te.columns]
            pred = model.predict(te[fc3].values)
            act  = te['ar'].values

            perf  = float(np.mean(((pred*act)>0)&(np.abs(act-pred)<=0.10)))
            trend = float(np.mean((pred*act)>0))
            n     = len(te)
            score = round(perf*0.4+trend*0.6,4)

            global_stats[hz_label]['perf'].append(perf)
            global_stats[hz_label]['trend'].append(trend)
            global_stats[hz_label]['n'].append(n)

            rec['acc_info'][hz_label]   = f"{perf:.1%}"
            rec['trend_info'][hz_label] = f"{trend:.1%}"
            rec['total_weight']        += n
            rec['final_score']          = max(rec['final_score'], score)

            ws_det.cell(row=det_row,column=1,value=hz_label)
            ws_det.cell(row=det_row,column=2,value=f"{perf:.1%}")
            ws_det.cell(row=det_row,column=3,value=f"{trend:.1%}")
            ws_det.cell(row=det_row,column=4,value=n)
            ws_det.cell(row=det_row,column=5,value=f"{pred.mean():.2%}")
            ws_det.cell(row=det_row,column=6,value=f"{act.mean():.2%}")
            ws_det.cell(row=det_row,column=7,value=f"{score:.3f}")
            print(f"   {hz_label}: 方向={trend:.1%}  完全={perf:.1%}  n={n}")
            det_row += 1

        auto_width(ws_det)
        results_summary.append(rec)

        # 計算整體 12M 勝率存入 power_data
        wr_12m = float(rec['trend_info'].get('12M','0%').rstrip('%'))/100
        if wr_12m >= 0.80:   stars='★★★★★'
        elif wr_12m >= 0.65: stars='★★★★☆'
        elif wr_12m >= 0.55: stars='★★★☆☆'
        elif wr_12m >= 0.45: stars='★★☆☆☆'
        else:                stars='★☆☆☆☆'
        tier = 'HIGH' if wr_12m>=0.65 else('MID' if wr_12m>=0.55 else 'LOW')
        power_data[code] = {'wr_12m':round(wr_12m,4),'tier':tier,
                             'stars':stars,'name':display,'category':category}

    # 填總覽
    results_summary.sort(key=lambda r: -float(r['trend_info'].get('12M','0%').rstrip('%')))
    POWER_COLOR = {'HIGH':'006100','MID':'7B5200','LOW':'9C0006'}
    POWER_BG    = {'HIGH':'C6EFCE','MID':'FFEB9C','LOW':'FFC7CE'}

    for ri, rec in enumerate(results_summary, 3):
        code  = rec['name'].split()[0]
        wr12  = float(rec['trend_info'].get('12M','0%').rstrip('%'))/100
        tf, _font, tl = tier_style(wr12)
        tier  = power_data.get(code,{}).get('tier','MID')
        stars = power_data.get(code,{}).get('stars','★★★☆☆')
        cat   = ETF_CATEGORY.get(code,'市值型')

        tier_label = {'HIGH':'強','MID':'中','LOW':'弱'}.get(tier,'中')
        ws_sum.cell(row=ri,column=1,value=rec['name']).font=Font(bold=True)
        ws_sum.cell(row=ri,column=2,value=cat)
        pc = ws_sum.cell(row=ri,column=3,value=f"{stars}  {tier_label}")
        pc.fill=PatternFill('solid',fgColor=POWER_BG.get(tier,'FFFFFF'))
        pc.font=Font(bold=True,color=POWER_COLOR.get(tier,'000000'))
        pc.alignment=Alignment(horizontal='center')
        ws_sum.cell(row=ri,column=4,value=f"{wr12:.1%}").alignment=Alignment(horizontal='center')
        col=5
        for hz in ['3M','6M','12M']:
            ws_sum.cell(row=ri,column=col,  value=rec['acc_info'].get(hz,'-'))
            tc=ws_sum.cell(row=ri,column=col+1,value=rec['trend_info'].get(hz,'-'))
            if rec['trend_info'].get(hz,'-')!='-':
                tv=float(rec['trend_info'][hz].rstrip('%'))/100
                ftf,_ff,_=tier_style(tv)
                tc.fill=ftf; tc.font=Font(bold=True)
            col+=2
        ws_sum.cell(row=ri,column=11,value=rec['total_weight'])
        ws_sum.cell(row=ri,column=12,value=rec['feat_count'])

    # 加權平均列
    cr = len(results_summary)+4
    cl = ws_sum.cell(row=cr,column=1,value="加權平均")
    cl.font=Font(bold=True,color=C_WHITE); cl.fill=PatternFill('solid',fgColor=C_NAVY)
    col=5
    for hz in ['3M','6M','12M']:
        ns=global_stats[hz]['n']
        if ns:
            wp=np.average(global_stats[hz]['perf'],weights=ns)
            wt=np.average(global_stats[hz]['trend'],weights=ns)
            ws_sum.cell(row=cr,column=col,  value=f"{wp:.1%}").font=Font(bold=True,color='FF0000')
            ws_sum.cell(row=cr,column=col+1,value=f"{wt:.1%}").font=Font(bold=True,color='FF0000')
        col+=2

    # 說明分頁
    ws_g = wb.create_sheet('📋 使用說明')
    guide = [
        ("【ETF 預測模型說明】","",True),
        ("模型","跨標的 XGBoost，用其他9支ETF的歷史訓練，預測目標ETF的方向",False),
        ("特徵","技術指標(11) + 總經指標(5) + ETF類型/Beta(5) + 估值動能(2)",False),
        ("預測週期","3M / 6M / 12M 方向（漲/跌）與幅度區間",False),
        ("","",False),
        ("【預測力說明】","",True),
        ("★★★★★","12M 方向勝率 ≥80%，訊號可靠性高",False),
        ("★★★★☆","12M 方向勝率 65~79%，訊號有一定參考性",False),
        ("★★★☆☆","12M 方向勝率 55~64%，訊號略優於隨機",False),
        ("★★☆☆☆","12M 方向勝率 45~54%，接近隨機，謹慎使用",False),
        ("★☆☆☆☆","12M 方向勝率 <45%，不建議參考",False),
        ("","",False),
        ("【使用注意事項】","",True),
        ("回測≠未來","歷史勝率不代表未來必然實現，ETF 受成分股調整、配息政策等影響",False),
        ("配息影響","高息型ETF除息後股價下跌屬正常，模型預測的是含息總報酬方向",False),
        ("新ETF樣本少","00762/00965等新上市ETF歷史資料較短，勝率參考性較低",False),
        ("建議搭配使用","模型訊號 + 大盤環境（VIX）+ 個人資金規劃，綜合判斷",False),
        ("定期更新","建議每季執行一次 etf_analysis.py + etf_backtester.py 更新資料",False),
    ]
    ws_g.column_dimensions['A'].width = 22
    ws_g.column_dimensions['B'].width = 75
    ws_g.merge_cells('A1:B1')
    _hdr(ws_g['A1'],'📋 ETF 預測模型使用說明',bg=C_NAVY,size=13)
    ws_g.row_dimensions[1].height = 28
    for i,(k,v,is_hdr) in enumerate(guide,2):
        if not k and not v: continue
        ka=ws_g.cell(row=i,column=1,value=k)
        va=ws_g.cell(row=i,column=2,value=v)
        if is_hdr:
            for cell in [ka,va]:
                cell.font=Font(bold=True,color=C_WHITE,size=11)
                cell.fill=PatternFill('solid',fgColor=C_BLUE)
        else:
            ka.font=Font(bold=True)
            va.alignment=Alignment(wrap_text=True,vertical='top')
        ws_g.row_dimensions[i].height = 36 if len(v)>50 else 18
        for cell in [ka,va]:
            cell.border=_thin()

    auto_width(ws_sum)
    wb.save(OUTPUT)
    print(f"\n✨ 完成！報告：{OUTPUT}")

    # 輸出 model_power.json
    with open('etf_model_power.json','w',encoding='utf-8') as f:
        json.dump(power_data, f, ensure_ascii=False, indent=2)
    print(f"📊 etf_model_power.json 已更新")

    return power_data

if __name__ == "__main__":
    run_backtest()
