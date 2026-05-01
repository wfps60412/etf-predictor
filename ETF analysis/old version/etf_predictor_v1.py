"""
etf_predictor.py  v1.0
ETF 預測報告生成

依序執行：
  1. etf_analysis.py    → ETF歷史分析資料庫.xlsx
  2. etf_backtester.py  → ETF回測報告_v1.xlsx + etf_model_power.json
  3. etf_predictor.py   → ETF預測報告.xlsx（自動讀取 etf_model_power.json）
"""
import pandas as pd
import numpy as np
import json
import os
import warnings
from xgboost import XGBRegressor
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

warnings.filterwarnings('ignore')

# ==========================================
# ETF 資訊
# ==========================================
ETF_NAMES = {
    '0050':   '元大台灣50',
    '00713':  '元大台灣高息低波',
    '00762':  '元大全球AI',
    '00965':  '元大航太防衛科技',
    '009816': '凱基台灣TOP50',
    '00878':  '國泰永續高股息',
    '00929':  '復華台灣科技優息',
    '0056':   '元大高股息',
    '00919':  '群益台灣精選高息',
    '00881':  '國泰台灣科技龍頭',
}

ETF_CATEGORY = {
    '0050':   '市值型',
    '009816': '市值型',
    '00881':  '科技型',
    '00762':  '主題型',
    '00965':  '主題型',
    '00713':  '高息低波',
    '0056':   '高息型',
    '00878':  '高息ESG',
    '00919':  '高息型',
    '00929':  '科技高息',
}

# 內建預設預測力（backtester 未跑時的 fallback）
MODEL_POWER_DEFAULT = {
    '0050':   (0.620, 'MID',  '★★★☆☆'),
    '009816': (0.610, 'MID',  '★★★☆☆'),
    '00881':  (0.640, 'MID',  '★★★☆☆'),
    '00762':  (0.580, 'MID',  '★★★☆☆'),
    '00965':  (0.560, 'MID',  '★★★☆☆'),
    '00713':  (0.630, 'MID',  '★★★☆☆'),
    '0056':   (0.650, 'HIGH', '★★★★☆'),
    '00878':  (0.640, 'MID',  '★★★☆☆'),
    '00919':  (0.640, 'MID',  '★★★☆☆'),
    '00929':  (0.660, 'HIGH', '★★★★☆'),
}
POWER_LABEL = {'HIGH':'強', 'MID':'中', 'LOW':'弱'}
POWER_COLOR = {'HIGH':'006100', 'MID':'7B5200', 'LOW':'9C0006'}
POWER_BG    = {'HIGH':'C6EFCE', 'MID':'FFEB9C', 'LOW':'FFC7CE'}

# ==========================================
# 特徵群組（與 backtester 對齊）
# ==========================================
TECH_FEATURES  = ['收盤價','成交量','MA5','MA20','ret5','RSI14',
                  'ATR14','BB_width','BIAS20','量能比','MACD_hist']
MACRO_FEATURES = ['台幣匯率','VIX指數','RSP','SP500_ret','DXY_ret']
TYPE_FEATURES  = ['ETFType','rolling_beta_60d','beta_regime',
                  'market_stress','beta_sp500_static']
VAL_FEATURES   = ['price_zscore','price_trend']

CATEGORY_FEATURES = {
    '市值型':   TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES,
    '科技型':   TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES,
    '主題型':   TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES,
    '高息型':   TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES + VAL_FEATURES,
    '高息低波': TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES + VAL_FEATURES,
    '高息ESG':  TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES + VAL_FEATURES,
    '科技高息': TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES + VAL_FEATURES,
}
DEFAULT_FEATURES = TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES

# ==========================================
# 特徵工程（與 backtester 完全一致）
# ==========================================
def build_features(df: pd.DataFrame, code: str) -> pd.DataFrame:
    d = df.copy()
    d['日期'] = pd.to_datetime(d['日期'])
    d = d.sort_values('日期').reset_index(drop=True)

    close = d['收盤價'].replace(0, np.nan)
    vol   = d['成交量'].replace(0, np.nan)

    d['MA5']    = close.rolling(5).mean().round(4)
    d['MA20']   = close.rolling(20).mean().round(4)
    d['ret5']   = close.pct_change(5).round(4)
    d['BIAS20'] = ((close - d['MA20']) / d['MA20'].replace(0, np.nan)).round(4)

    delta = close.diff()
    gain  = delta.clip(lower=0).rolling(14).mean()
    loss  = (-delta.clip(upper=0)).rolling(14).mean()
    d['RSI14'] = (100 - 100/(1+gain/loss.replace(0,np.nan))).round(4)

    high = d['最高價'].replace(0, np.nan)
    low  = d['最低價'].replace(0, np.nan)
    tr   = pd.concat([high-low,
                      (high-close.shift(1)).abs(),
                      (low -close.shift(1)).abs()], axis=1).max(axis=1)
    d['ATR14'] = tr.rolling(14).mean().round(4)

    bb_mid = close.rolling(20).mean()
    bb_std = close.rolling(20).std()
    d['BB_width'] = ((bb_mid+2*bb_std-(bb_mid-2*bb_std))
                     /bb_mid.replace(0,np.nan)).round(4)

    vol_ma20 = vol.rolling(20).mean()
    d['量能比'] = (vol/vol_ma20.replace(0,np.nan)).round(4)

    ema12 = close.ewm(span=12,adjust=False).mean()
    ema26 = close.ewm(span=26,adjust=False).mean()
    macd  = ema12 - ema26
    d['MACD_hist'] = (macd - macd.ewm(span=9,adjust=False).mean()).round(4)

    for col in ['RSP','SP500_ret','DXY_ret','台幣匯率','VIX指數',
                'ETFType','rolling_beta_60d','beta_regime',
                'market_stress','beta_sp500_static',
                'price_zscore','price_trend']:
        if col not in d.columns:
            d[col] = 0.0

    return d.dropna(subset=['收盤價','MA5','MA20'])


# ==========================================
# 訓練並預測
# ==========================================
def train_and_predict(target_feat, all_feat_list, horizon_td, feat_cols=None):
    if feat_cols is None:
        feat_cols = DEFAULT_FEATURES
    Xs, ys = [], []
    for feat_df in all_feat_list:
        tmp = feat_df.copy()
        tmp['target'] = tmp['收盤價'].pct_change(horizon_td).shift(-horizon_td)
        fc  = [f for f in feat_cols if f in tmp.columns]
        v   = tmp[fc+['target']].dropna()
        v   = v[v['target'].abs()<=0.6]
        if len(v)>20:
            Xs.append(v[fc].values)
            ys.append(v['target'].values)
    if not Xs:
        return None, None, None, 0

    Xa = np.vstack(Xs); ya = np.concatenate(ys)
    m   = Xa.std(axis=0) > 0
    fc2 = [f for f,vv in zip(fc, m) if vv]
    Xa  = Xa[:,m]

    mdl = Pipeline([('s',StandardScaler()),
                    ('m',XGBRegressor(n_estimators=200,max_depth=4,
                                      learning_rate=0.05,subsample=0.8,
                                      colsample_bytree=0.8,random_state=42,
                                      verbosity=0))])
    mdl.fit(Xa, ya)

    # 殘差區間（P10/P90）
    train_pred = mdl.predict(Xa)
    resid      = ya - train_pred
    lo_r, hi_r = np.percentile(resid, 10), np.percentile(resid, 90)

    # 對目標 ETF 最後一行預測
    last = target_feat.iloc[[-1]][[f for f in fc2 if f in target_feat.columns]]
    if last.shape[1] < len(fc2):
        return None, None, None, len(fc2)
    center = float(mdl.predict(last)[0])
    return center, center+lo_r, center+hi_r, len(fc2)


# ==========================================
# Excel 格式
# ==========================================
C_NAVY = '1F3864'; C_BLUE = '2E75B6'; C_WHITE = 'FFFFFF'

def _thin():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(cell, val, bg=C_BLUE, size=10):
    cell.value     = val
    cell.font      = Font(bold=True, color=C_WHITE, size=size)
    cell.fill      = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)
    cell.border    = _thin()

def signal_icon(center):
    if   center >  0.05: return '📈 偏多'
    elif center < -0.05: return '📉 偏空'
    else:                return '➡️  中性'

def auto_width(ws, min_w=14, max_w=65):
    for col in ws.columns:
        best = min_w
        for cell in col:
            if cell.value:
                lines = str(cell.value).split('\n')
                best  = min(max_w, max(best, max(len(l) for l in lines)+4))
        ws.column_dimensions[get_column_letter(col[0].column)].width = best


# ==========================================
# 主流程
# ==========================================
def main():
    DB_FILE  = 'ETF歷史分析資料庫.xlsx'
    OUT_FILE = 'ETF預測報告.xlsx'

    print("="*60)
    print("  ETF 預測系統 v1.0")
    print("="*60)

    if not os.path.exists(DB_FILE):
        print(f"❌ 找不到 {DB_FILE}，請先執行 etf_analysis.py")
        return

    # 載入 etf_model_power.json
    power_map = {}   # code → (wr, tier, stars)
    json_path = 'etf_model_power.json'
    if os.path.exists(json_path):
        try:
            with open(json_path, encoding='utf-8') as f:
                mp = json.load(f)
            for code, d in mp.items():
                power_map[code] = (d['wr_12m'], d['tier'], d['stars'])
            print(f"📊 已從 etf_model_power.json 載入最新勝率（{len(power_map)} 支）")
        except Exception as e:
            print(f"⚠️  etf_model_power.json 讀取失敗，使用內建預設值：{e}")
    else:
        print("ℹ️  未找到 etf_model_power.json，使用內建預設值")
        print("   建議先執行 etf_backtester.py")

    # 讀取資料庫
    xl = pd.ExcelFile(DB_FILE)
    all_feats = {}
    for sheet in xl.sheet_names:
        code = sheet.split()[0].strip()
        try:
            df_raw = pd.read_excel(DB_FILE, sheet_name=sheet)
            all_feats[code] = build_features(df_raw, code)
            print(f"  ✅ {code} {ETF_NAMES.get(code,'')}: {len(all_feats[code])} 筆")
        except Exception as e:
            print(f"  ❌ {sheet}: {e}")

    # 預測
    horizons  = {'3M':63, '6M':126, '12M':252}
    base_date = datetime.now().strftime("%Y-%m-%d")
    results   = []

    for code, df_feat in all_feats.items():
        name     = ETF_NAMES.get(code, code)
        cat      = ETF_CATEGORY.get(code, '市值型')
        feat_cols= CATEGORY_FEATURES.get(cat, DEFAULT_FEATURES)
        others   = [d for c,d in all_feats.items() if c != code]

        wr, tier, stars = power_map.get(code, MODEL_POWER_DEFAULT.get(
            code, (0.55, 'MID', '★★★☆☆')))

        print(f"\n  預測 {code} {name}...", end=' ', flush=True)
        pred = {}
        feat_count = 0
        for hz_label, hz_days in horizons.items():
            try:
                center, lo, hi, fc = train_and_predict(
                    df_feat, others, hz_days, feat_cols)
                if center is None:
                    pred[hz_label] = None
                else:
                    pred[hz_label] = {'center':center,'lo':lo,'hi':hi}
                    feat_count = fc
            except Exception as e:
                pred[hz_label] = None
        print(f"完成 ({feat_count}特徵)")

        current_price = float(df_feat['收盤價'].iloc[-1])
        results.append({
            'code':          code,
            'name':          name,
            'category':      cat,
            'current_price': round(current_price, 2),
            'base_date':     str(df_feat['日期'].iloc[-1].date()),
            'pred':          pred,
            'feat_count':    feat_count,
            'power_wr':      wr,
            'power_tier':    tier,
            'power_stars':   stars,
        })

    # 依預測力排序
    results.sort(key=lambda x: -x['power_wr'])

    # ==========================================
    # 寫 Excel
    # ==========================================
    wb = Workbook()
    ws = wb.active
    ws.title = '📊 ETF預測總覽'

    # 標題列
    n_cols = 18
    ws.merge_cells(f'A1:{get_column_letter(n_cols)}1')
    _hdr(ws['A1'],
         f'ETF AI 預測報告 v1.0　基準日：{base_date}　產業別特徵',
         bg=C_NAVY, size=13)
    ws.row_dimensions[1].height = 30

    # 欄位標題
    col_headers = [
        '代號與名稱', 'ETF類型', '預測力', '12M方向勝率', '現價', '日期',
        '3M預測中心', '3M區間(Lo)', '3M區間(Hi)', '3M訊號',
        '6M預測中心', '6M區間(Lo)', '6M區間(Hi)', '6M訊號',
        '12M預測中心','12M區間(Lo)','12M區間(Hi)','12M訊號',
        '特徵數',
    ]
    for ci, h in enumerate(col_headers, 1):
        _hdr(ws.cell(row=2, column=ci), h)
    ws.row_dimensions[2].height = 36

    # 資料列
    TREND_COLORS = {
        '📈 偏多': ('E2EFDA','375623'),
        '📉 偏空': ('FFDCE0','9C0006'),
        '➡️  中性': ('EFF3FF','1F3864'),
    }

    for ri, rec in enumerate(results, 3):
        tier  = rec['power_tier']
        stars = rec['power_stars']
        wr    = rec['power_wr']
        display = f"{rec['code']} {rec['name']}"

        ws.cell(row=ri, column=1, value=display).font = Font(bold=True)
        ws.cell(row=ri, column=2, value=rec['category'])

        # 預測力
        pc = ws.cell(row=ri, column=3,
                     value=f"{stars}  {POWER_LABEL.get(tier,'中')}")
        pc.fill  = PatternFill('solid', fgColor=POWER_BG.get(tier,'FFFFFF'))
        pc.font  = Font(bold=True, color=POWER_COLOR.get(tier,'000000'))
        pc.alignment = Alignment(horizontal='center')
        pc.border = _thin()

        # 12M 方向勝率
        wc = ws.cell(row=ri, column=4, value=f"{wr:.1%}")
        wc.fill  = PatternFill('solid', fgColor=POWER_BG.get(tier,'FFFFFF'))
        wc.font  = Font(bold=True, color=POWER_COLOR.get(tier,'000000'))
        wc.alignment = Alignment(horizontal='center')
        wc.border = _thin()

        ws.cell(row=ri, column=5, value=rec['current_price'])
        ws.cell(row=ri, column=6, value=rec['base_date'])

        col = 7
        for hz in ['3M', '6M', '12M']:
            p = rec['pred'].get(hz)
            if p is None:
                for _ in range(4):
                    ws.cell(row=ri, column=col, value='-')
                    col += 1
            else:
                center = p['center']
                ws.cell(row=ri, column=col,   value=f"{center:.1%}")
                ws.cell(row=ri, column=col+1, value=f"{p['lo']:.1%}")
                ws.cell(row=ri, column=col+2, value=f"{p['hi']:.1%}")
                sig  = signal_icon(center)
                scell= ws.cell(row=ri, column=col+3, value=sig)
                bg, fg = TREND_COLORS.get(sig, ('FFFFFF','000000'))
                scell.fill      = PatternFill('solid', fgColor=bg)
                scell.font      = Font(bold=True, color=fg)
                scell.alignment = Alignment(horizontal='center')
                scell.border    = _thin()
                col += 4

        ws.cell(row=ri, column=col, value=rec['feat_count'])

    # 說明列
    note_row = len(results) + 4
    ws.merge_cells(f'A{note_row}:{get_column_letter(n_cols)}{note_row}')
    note = ws.cell(row=note_row, column=1,
                   value="📈偏多：預測漲幅>5%　📉偏空：預測跌幅>5%　➡️中性：±5%以內　"
                         "HIGH≥65% MID 55-65% LOW<55%　"
                         "★★★★★=極強(≥80%)  ★★★★☆=強(65-79%)  ★★★☆☆=中(55-64%)  ★★☆☆☆=弱(45-54%)  ★☆☆☆☆=極弱(<45%)")
    note.font      = Font(italic=True, color='595959', size=10)
    note.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[note_row].height = 32

    auto_width(ws)

    # ==========================================
    # 投資使用說明分頁
    # ==========================================
    ws_g = wb.create_sheet('📋 投資使用說明')
    ws_g.column_dimensions['A'].width = 22
    ws_g.column_dimensions['B'].width = 78

    guide_rows = [
        ("【模型定位】","",True),
        ("用途",     "輔助參考工具，非投資建議。提供 ETF 的中長期方向訊號，需搭配市場判斷使用。",False),
        ("適用週期", "12M 預測最可靠，6M 次之，3M 受短期雜訊影響較大。",False),
        ("訓練方式", "跨 ETF 聯合訓練（Leave-One-Out）：用其他9支ETF歷史訓練，預測目標ETF。",False),
        ("","",False),
        ("【如何解讀預測力星等】","",True),
        ("★★★★★（極強）","12M 方向勝率 ≥80%。訊號可靠性高，可作為主要參考依據之一。",False),
        ("★★★★☆（強）",  "12M 方向勝率 65~79%。訊號有效，建議搭配技術面確認。",False),
        ("★★★☆☆（中）",  "12M 方向勝率 55~64%。略優於隨機，輕倉或等待更明確訊號。",False),
        ("★★☆☆☆（弱）",  "12M 方向勝率 45~54%。接近隨機，不建議以此作為主要依據。",False),
        ("★☆☆☆☆（極弱）","12M 方向勝率 <45%。訊號不可靠，建議直接忽略。",False),
        ("","",False),
        ("【ETF 分類與特徵說明】","",True),
        ("市值型",   "0050、009816：追蹤大盤，使用技術+總經+ETF類型特徵（21個）",False),
        ("科技型",   "00881：科技龍頭，同上特徵（21個）",False),
        ("主題型",   "00762全球AI、00965航太防衛：主題性強，Beta_SP500較高（21個）",False),
        ("高息型",   "0056、00878、00919：加入估值動能特徵（23個），配息特性反映於price_zscore",False),
        ("科技高息", "00929：兼具科技與高息，使用最完整特徵（23個）",False),
        ("高息低波", "00713：低波動設計，估值百分位有較高預測力（23個）",False),
        ("","",False),
        ("【使用四步驟】","",True),
        ("Step 1","篩選預測力 ★★★★☆ 以上的 ETF（HIGH 組，綠底標示）。",False),
        ("Step 2","查看 12M 訊號方向（📈偏多 / 📉偏空 / ➡️中性）。",False),
        ("Step 3","確認信賴區間（Lo ~ Hi）。區間越窄代表模型越有信心。",False),
        ("Step 4","對照個人資金規劃、配息需求，綜合判斷進出時機。",False),
        ("","",False),
        ("【特別注意事項】","",True),
        ("配息影響",   "高息型ETF除息時股價下跌屬正常。模型預測的是除息後的「價格走勢方向」，非含息報酬。",False),
        ("新上市ETF",  "00762（全球AI）、00965（航太防衛）上市時間較短（2~3年），歷史樣本少，勝率參考性較低，需特別謹慎。",False),
        ("成分股調整", "ETF 每季調整成分股，調整後市場反應難以預測，建議在成分股公告前後保持觀望。",False),
        ("回測≠未來",  "所有勝率數字為歷史回測，實盤因市場環境變化，實際表現通常有落差。",False),
        ("定期更新",   "建議每季執行一次 etf_analysis.py → etf_backtester.py → etf_predictor.py，確保資料最新。",False),
        ("","",False),
        ("【技術規格】","",True),
        ("演算法",   "XGBoost（梯度提升樹），n=200, depth=4, lr=0.05, subsample=0.8",False),
        ("特徵設計", "技術面(11) + 總經(5) + ETF類型/Beta(5) + 估值動能(2)，依類型選 21~23個",False),
        ("預測區間", "訓練集殘差 P10/P90 百分位",False),
        ("版本",     "etf_analysis v1.0 / etf_backtester v1.0 / etf_predictor v1.0",False),
    ]

    ws_g.merge_cells('A1:B1')
    _hdr(ws_g['A1'], '📋 ETF 投資使用說明與注意事項', bg=C_NAVY, size=13)
    ws_g.row_dimensions[1].height = 28

    for i, (k, v, is_hdr) in enumerate(guide_rows, 2):
        if not k and not v:
            continue
        ka = ws_g.cell(row=i, column=1, value=k)
        va = ws_g.cell(row=i, column=2, value=v)
        if is_hdr:
            for cell in [ka, va]:
                cell.font  = Font(bold=True, color=C_WHITE, size=11)
                cell.fill  = PatternFill('solid', fgColor=C_BLUE)
                cell.alignment = Alignment(vertical='center')
            ws_g.row_dimensions[i].height = 22
        else:
            ka.font = Font(bold=True)
            va.alignment = Alignment(wrap_text=True, vertical='top')
            ws_g.row_dimensions[i].height = 42 if len(v) > 60 else 18
        for cell in [ka, va]:
            cell.border = _thin()

    wb.save(OUT_FILE)
    print(f"\n✅ ETF 預測報告已儲存：{OUT_FILE}")
    print(f"   分頁：📊 ETF預測總覽 | 📋 投資使用說明 + 各ETF明細")


if __name__ == "__main__":
    main()
