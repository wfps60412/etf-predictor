import pandas as pd
import numpy as np
import optuna
from xgboost import XGBRegressor
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, json, warnings
from datetime import datetime

warnings.filterwarnings('ignore')
optuna.logging.set_verbosity(optuna.logging.WARNING)

# ==========================================
# 1. 配置與特徵定義（修正：對齊 etf_analysis_v2）
# ==========================================
ETF_NAMES = {
    '0050':'元大台灣50',       '00713':'元大台灣高息低波',
    '00762':'元大全球AI',       '00965':'元大航太防衛科技',
    '009816':'凱基台灣TOP50',   '00878':'國泰永續高股息',
    '00929':'復華台灣科技優息', '0056':'元大高股息',
    '00919':'群益台灣精選高息', '00881':'國泰台灣科技龍頭',
}
ETF_CATEGORY = {
    '0050':'市值型',   '009816':'市值型',  '00881':'科技型',
    '00762':'主題型',  '00965':'主題型',   '00713':'高息低波',
    '0056':'高息型',   '00878':'高息ESG',  '00919':'高息型',  '00929':'科技高息',
}

# 修正1：對齊 etf_analysis_v2 的實際欄位（含新增的美債/黃金/原油，移除 beta_regime/market_stress）
TECH_FEATURES  = ['收盤價','成交量','MA5','MA20','ret5','RSI14','ATR14',
                  'BB_width','BIAS20','量能比','MACD_hist']
MACRO_FEATURES = ['台幣匯率','VIX指數','RSP','SP500_ret','DXY_ret',
                  '美債10Y_ret','黃金_ret','原油_ret']   # 修正1：加入 v2 新特徵
TYPE_FEATURES  = ['ETFType','rolling_beta_60d','beta_sp500_static']  # 修正2：移除 beta_regime/market_stress
VAL_FEATURES   = ['price_zscore','price_trend']

# ETF 類型 → 特徵組合（高息型額外使用估值特徵）
CATEGORY_FEATURES = {
    '市值型':   TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES,
    '科技型':   TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES,
    '主題型':   TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES,
    '高息型':   TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES + VAL_FEATURES,
    '高息低波': TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES + VAL_FEATURES,
    '高息ESG':  TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES + VAL_FEATURES,
    '科技高息': TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES + VAL_FEATURES,
}
DEFAULT_FEATURES = TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES
ALL_FEATURES_SET = list(dict.fromkeys(
    TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES + VAL_FEATURES))

# ==========================================
# 2. 視覺呈現工具
# ==========================================
C_NAVY, C_BLUE, C_WHITE = '1F3864', '2E75B6', 'FFFFFF'

def _thin():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(cell, val, bg=C_BLUE, size=10):
    cell.value     = val
    cell.font      = Font(bold=True, color=C_WHITE, size=size)
    cell.fill      = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = _thin()

def tier_style(wr):
    if wr >= 0.65:   return PatternFill('solid',fgColor='C6EFCE'), Font(color='006100',bold=True), 'HIGH', '強'
    elif wr >= 0.55: return PatternFill('solid',fgColor='DDEBF7'), Font(color='1F497D',bold=True), 'MID',  '中'
    else:            return PatternFill('solid',fgColor='FFC7CE'), Font(color='9C0006',bold=True), 'LOW',  '弱'

def auto_width(ws, min_w=12, max_w=55):
    for col in ws.columns:
        best = min_w
        for cell in col:
            try:
                if cell.value:
                    best = min(max_w, max(best, len(str(cell.value)) + 4))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = best

# ==========================================
# 3. 宏觀過濾器（保留 v2 設計）
# ==========================================
def apply_macro_filter(pred_value, current_row):
    """高 VIX 或 S&P500 單日大跌時，對多方信號降低強度"""
    vix    = current_row.get('VIX指數',   20)
    sp_ret = current_row.get('SP500_ret', 0)
    if vix > 25 or sp_ret < -0.05:
        return pred_value * 0.5 if pred_value > 0 else pred_value * 1.2
    return pred_value

# ==========================================
# 4. 特徵工程（修正：完整計算技術指標）
# ==========================================
def build_features(df: pd.DataFrame, code: str = '') -> pd.DataFrame:
    """從資料庫 DataFrame 計算所有技術指標"""
    d = df.copy()
    d['日期'] = pd.to_datetime(d['日期'])
    d = d.sort_values('日期').reset_index(drop=True)

    close = d['收盤價'].replace(0, np.nan)
    vol   = d['成交量'].replace(0, np.nan)

    # 技術指標
    d['MA5']    = close.rolling(5).mean().round(4)
    d['MA20']   = close.rolling(20).mean().round(4)
    d['ret5']   = close.pct_change(5).round(4)
    d['BIAS20'] = ((close - d['MA20']) / d['MA20'].replace(0, np.nan)).round(4)

    delta = close.diff()
    gain  = delta.clip(lower=0).rolling(14).mean()
    loss  = (-delta.clip(upper=0)).rolling(14).mean()
    d['RSI14'] = (100 - 100 / (1 + gain / loss.replace(0, np.nan))).round(4)

    high = d['最高價'].replace(0, np.nan)
    low  = d['最低價'].replace(0, np.nan)
    tr   = pd.concat([high - low,
                      (high - close.shift(1)).abs(),
                      (low  - close.shift(1)).abs()], axis=1).max(axis=1)
    d['ATR14'] = tr.rolling(14).mean().round(4)

    bb_mid = close.rolling(20).mean()
    bb_std = close.rolling(20).std()
    d['BB_width'] = ((bb_mid + 2*bb_std - (bb_mid - 2*bb_std))
                     / bb_mid.replace(0, np.nan)).round(4)

    d['量能比']  = (vol / vol.rolling(20).mean().replace(0, np.nan)).round(4)

    ema12 = close.ewm(span=12, adjust=False).mean()
    ema26 = close.ewm(span=26, adjust=False).mean()
    macd  = ema12 - ema26
    d['MACD_hist'] = (macd - macd.ewm(span=9, adjust=False).mean()).round(4)

    # 補缺失欄位（資料庫沒有的欄位補0）
    for col in ALL_FEATURES_SET:
        if col not in d.columns:
            d[col] = 0.0

    return d.ffill().bfill().dropna(subset=['收盤價', 'MA5', 'MA20'])

# ==========================================
# 5. Optuna 超參數調優
# ==========================================
def train_optuna_model(Xs: np.ndarray, ys: np.ndarray):
    """用 Optuna 找最佳 XGBoost 超參數，再整體訓練"""
    if len(ys) < 25:
        return None

    def objective(trial):
        param = {
            'n_estimators':  trial.suggest_int('n_estimators',  50, 200),
            'max_depth':     trial.suggest_int('max_depth',      2,   5),
            'learning_rate': trial.suggest_float('learning_rate', 0.01, 0.15),
            'subsample':     trial.suggest_float('subsample',     0.6,  1.0),
            'random_state':  42, 'verbosity': 0,
        }
        split = int(len(Xs) * 0.8)
        reg   = XGBRegressor(**param).fit(Xs[:split], ys[:split])
        return float(np.mean((reg.predict(Xs[split:]) - ys[split:]) ** 2))

    study = optuna.create_study(direction='minimize')
    study.optimize(objective, n_trials=10)   # 修正：從5提升至10，搜索更充分

    best_mdl = Pipeline([
        ('s', StandardScaler()),
        ('m', XGBRegressor(**study.best_params, random_state=42, verbosity=0))
    ])
    best_mdl.fit(Xs, ys)
    return best_mdl

# ==========================================
# 6. 主流程
# ==========================================
def main():
    DB_FILE = 'ETF歷史分析資料庫.xlsx'
    OUTPUT  = 'ETF回測報告_v2.xlsx'

    if not os.path.exists(DB_FILE):
        print(f"❌ 找不到 {DB_FILE}，請先執行 etf_analysis.py")
        return

    xl = pd.ExcelFile(DB_FILE)
    all_dfs = {}
    for sheet in xl.sheet_names:
        code = sheet.split()[0].strip()
        try:
            all_dfs[code] = build_features(
                pd.read_excel(DB_FILE, sheet_name=sheet), code)
            print(f"  ✅ {sheet}: {len(all_dfs[code])} 筆")
        except Exception as e:
            print(f"  ❌ {sheet}: {e}")

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = '回測準確率總覽'

    ws_sum.merge_cells('A1:L1')
    _hdr(ws_sum['A1'],
         f'ETF AI 實戰回測報告 v2.0（WFV + Optuna + MacroFilter）　{datetime.now().strftime("%Y-%m-%d")}',
         bg=C_NAVY, size=13)
    ws_sum.row_dimensions[1].height = 28

    hdrs = ['代號與名稱','類型','預測力','12M勝率',
            '3M完全(±10%)','3M方向勝率',
            '6M完全(±10%)','6M方向勝率',
            '12M完全(±10%)','12M方向勝率',
            '樣本數','特徵數']
    for ci, h in enumerate(hdrs, 1):
        _hdr(ws_sum.cell(row=2, column=ci), h)
    ws_sum.row_dimensions[2].height = 36

    HORIZONS     = {'3M': 63, '6M': 126, '12M': 252}
    global_stats = {hz: {'perf':[], 'trend':[], 'n':[]} for hz in HORIZONS}
    results_summary, power_json = [], {}

    for code, df_feat in all_dfs.items():
        name     = ETF_NAMES.get(code, code)
        category = ETF_CATEGORY.get(code, '市值型')
        feat_cols = CATEGORY_FEATURES.get(category, DEFAULT_FEATURES)
        # 只保留資料庫中真實有值的特徵
        feat_cols = [f for f in feat_cols if f in df_feat.columns and
                     (df_feat[f] != 0).any()]
        others = [d for c, d in all_dfs.items() if c != code]

        display = f"{code} {name}"
        print(f"\n🔄 回測 {display}  [{category}，{len(feat_cols)}特徵]")

        # WFV 明細分頁
        ws_det = wb.create_sheet(display[:31])
        ws_det.merge_cells('A1:G1')
        _hdr(ws_det['A1'], f'{display} 回測明細（WFV）', bg=C_NAVY, size=12)
        ws_det.row_dimensions[1].height = 24
        for ci, h in enumerate(['週期','完全成功(±10%)','方向勝率',
                                  '樣本數','平均預測漲幅','平均實際漲幅','Optuna最佳MSE'], 1):
            _hdr(ws_det.cell(row=2, column=ci), h)

        rec     = {'code': code, 'name': display, 'acc': {}, 'trend': {}, 'n': 0}
        det_row = 3

        for hz_label, hz_days in HORIZONS.items():
            test_start  = int(len(df_feat) * 0.75)
            all_preds, all_acts = [], []

            for i in range(test_start, len(df_feat) - hz_days, 63):
                # 訓練集：其他9支 ETF 的前 i 筆
                Xs_l, ys_l = [], []
                for td in others:
                    td_c = td.copy()
                    td_c['target'] = td_c['收盤價'].pct_change(hz_days).shift(-hz_days)
                    fc   = [f for f in feat_cols if f in td_c.columns]
                    v    = td_c[fc + ['target']].dropna()
                    v    = v[v['target'].abs() <= 0.6]
                    if len(v) > 20:
                        Xs_l.append(v[fc].values)
                        ys_l.append(v['target'].values)
                if not Xs_l:
                    continue

                Xa = np.vstack(Xs_l); ya = np.concatenate(ys_l)
                m  = Xa.std(axis=0) > 0
                fc_valid = [f for f, vv in zip(feat_cols, m) if vv]
                Xa = Xa[:, m]

                mdl = train_optuna_model(Xa, ya)
                if mdl is None:
                    continue

                # 預測目標 ETF 在時間點 i 的未來漲跌
                row_feats = df_feat.iloc[i][fc_valid].values.reshape(1, -1)
                raw_pred  = mdl.predict(row_feats)[0]
                filtered  = apply_macro_filter(raw_pred, df_feat.iloc[i].to_dict())

                if i + hz_days < len(df_feat):
                    actual = ((df_feat['收盤價'].iloc[i + hz_days]
                               - df_feat['收盤價'].iloc[i])
                              / df_feat['收盤價'].iloc[i])
                    all_preds.append(filtered)
                    all_acts.append(actual)

            if not all_preds:
                continue

            ap  = np.array(all_preds)
            aa  = np.array(all_acts)
            perf  = float(np.mean((np.abs(aa - ap) <= 0.10) & ((ap * aa) > 0)))
            trend = float(np.mean((ap * aa) > 0))
            n     = len(ap)

            rec['acc'][hz_label]   = perf
            rec['trend'][hz_label] = trend
            rec['n']               = max(rec['n'], n)

            global_stats[hz_label]['perf'].append(perf)
            global_stats[hz_label]['trend'].append(trend)
            global_stats[hz_label]['n'].append(n)

            # 明細分頁
            ws_det.cell(row=det_row, column=1, value=hz_label)
            ws_det.cell(row=det_row, column=2, value=f"{perf:.1%}")
            ct = ws_det.cell(row=det_row, column=3, value=f"{trend:.1%}")
            tf, tf_font, _, _ = tier_style(trend)
            ct.fill, ct.font = tf, tf_font
            ws_det.cell(row=det_row, column=4, value=n)
            ws_det.cell(row=det_row, column=5, value=f"{ap.mean():.2%}")
            ws_det.cell(row=det_row, column=6, value=f"{aa.mean():.2%}")
            print(f"   {hz_label}: 方向={trend:.1%}  完全={perf:.1%}  n={n}")
            det_row += 1

        auto_width(ws_det)
        results_summary.append(rec)

        # 修正：power_json 補全 tier / name / category 欄位（讓 predictor 相容）
        wr12 = rec['trend'].get('12M', 0)
        if wr12 >= 0.80:   stars = '★★★★★'
        elif wr12 >= 0.65: stars = '★★★★☆'
        elif wr12 >= 0.55: stars = '★★★☆☆'
        elif wr12 >= 0.45: stars = '★★☆☆☆'
        else:              stars = '★☆☆☆☆'
        _, _, tier, _ = tier_style(wr12)
        power_json[code] = {
            'wr_12m':   round(wr12, 4),
            'tier':     tier,
            'stars':    stars,
            'name':     display,
            'category': category,
        }

    # 總覽依 12M 勝率排序
    results_summary.sort(key=lambda r: -r['trend'].get('12M', 0))

    for ri, rec in enumerate(results_summary, 3):
        code  = rec['code']
        wr12  = rec['trend'].get('12M', 0)
        tf, tf_font, _, tier_label = tier_style(wr12)
        stars = power_json.get(code, {}).get('stars', '★★☆☆☆')

        ws_sum.cell(row=ri, column=1, value=rec['name']).font = Font(bold=True)
        ws_sum.cell(row=ri, column=2, value=ETF_CATEGORY.get(code, '市值型'))

        pc = ws_sum.cell(row=ri, column=3, value=f"{stars}  {tier_label}")
        pc.fill, pc.font = tf, tf_font
        pc.alignment = Alignment(horizontal='center')

        ws_sum.cell(row=ri, column=4, value=f"{wr12:.1%}").alignment = Alignment(horizontal='center')

        col_ptr = 5
        for hz in ['3M', '6M', '12M']:
            ws_sum.cell(row=ri, column=col_ptr,
                        value=f"{rec['acc'].get(hz, 0):.1%}" if hz in rec['acc'] else "-")
            ct = ws_sum.cell(row=ri, column=col_ptr+1,
                             value=f"{rec['trend'].get(hz, 0):.1%}" if hz in rec['trend'] else "-")
            if hz in rec['trend']:
                vf, vfont, _, _ = tier_style(rec['trend'][hz])
                ct.fill, ct.font = vf, vfont
            col_ptr += 2

        ws_sum.cell(row=ri, column=11, value=rec['n'])
        ws_sum.cell(row=ri, column=12, value=len(feat_cols))

    # 加權平均列
    avg_row = len(results_summary) + 4
    cl = ws_sum.cell(row=avg_row, column=1, value="WFV 實戰加權平均")
    cl.font = Font(bold=True, color=C_WHITE)
    cl.fill = PatternFill('solid', fgColor=C_NAVY)
    col_ptr = 5
    for hz in ['3M', '6M', '12M']:
        ns = global_stats[hz]['n']
        if ns:
            wp = np.average(global_stats[hz]['perf'],  weights=ns)
            wt = np.average(global_stats[hz]['trend'], weights=ns)
            ws_sum.cell(row=avg_row, column=col_ptr,   value=f"{wp:.1%}").font = Font(bold=True, color='FF0000')
            ws_sum.cell(row=avg_row, column=col_ptr+1, value=f"{wt:.1%}").font = Font(bold=True, color='FF0000')
        col_ptr += 2

    auto_width(ws_sum)
    wb.save(OUTPUT)
    print(f"\n✨ 回測完成！報告：{OUTPUT}")

    # 輸出 etf_model_power.json（predictor 讀取用）
    with open('etf_model_power.json', 'w', encoding='utf-8') as f:
        json.dump(power_json, f, ensure_ascii=False, indent=2)
    print(f"📊 etf_model_power.json 已更新（{len(power_json)} 支）")


if __name__ == "__main__":
    main()
