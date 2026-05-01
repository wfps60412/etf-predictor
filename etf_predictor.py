"""
etf_predictor.py v2.1
ETF AI 預測 - 真實模型版

特徵集對齊 etf_analysis v2.1：
  技術面(11) + 宏觀面(8，含美債/黃金/原油) + 類型面(3) + 估值面(2)
預測力星等從 etf_model_power.json 讀取（回測產出）
跨 ETF 聯合訓練 + Optuna 超參數調優 + 動態信賴區間
"""
import os, sys, json, warnings
try:
    from tqdm import tqdm
except ImportError:
    def tqdm(iterable, **kw):
        items = list(iterable)
        n = len(items)
        desc = kw.get('desc', '')
        for i, item in enumerate(items):
            pct = (i + 1) / n
            bar = '█' * int(pct * 25) + '░' * (25 - int(pct * 25))
            print(f'\r  {desc} [{bar}] {i+1}/{n} ({pct:.0%})', end='', flush=True)
            yield item
        print()
import numpy as np
import pandas as pd
import optuna
from xgboost import XGBRegressor
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

warnings.filterwarnings('ignore')
optuna.logging.set_verbosity(optuna.logging.WARNING)

# ==========================================
# 1. 配置
# ==========================================
ETF_NAMES = {
    '0050':'元大台灣50',       '00713':'元大台灣高息低波',
    '00762':'元大全球AI',       '00965':'元大航太防衛科技',
    '009816':'凱基台灣TOP50',   '00878':'國泰永續高股息',
    '00929':'復華台灣科技優息', '0056':'元大高股息',
    '00919':'群益台灣精選高息', '00881':'國泰台灣科技龍頭',
}
ETF_CATEGORY = {
    '0050':'市值型',   '009816':'市值型',  '00881':'科技型',
    '00762':'主題型',  '00965':'主題型',   '00713':'高息低波',
    '0056':'高息型',   '00878':'高息ESG',  '00919':'高息型',  '00929':'科技高息',
}

TECH_FEATURES  = ['收盤價','成交量','MA5','MA20','ret5','RSI14','ATR14',
                  'BB_width','BIAS20','量能比','MACD_hist']
MACRO_FEATURES = ['台幣匯率','VIX指數','RSP','SP500_ret','DXY_ret',
                  '美債10Y_ret','黃金_ret','原油_ret']
TYPE_FEATURES  = ['ETFType','rolling_beta_60d','beta_sp500_static']
VAL_FEATURES   = ['price_zscore','price_trend']

CATEGORY_FEATURES = {
    '市值型':   TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES,
    '科技型':   TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES,
    '主題型':   TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES,
    '高息型':   TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES + VAL_FEATURES,
    '高息低波': TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES + VAL_FEATURES,
    '高息ESG':  TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES + VAL_FEATURES,
    '科技高息': TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES + VAL_FEATURES,
}
DEFAULT_FEATURES = TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES
ALL_FEATURES_SET = list(dict.fromkeys(
    TECH_FEATURES + MACRO_FEATURES + TYPE_FEATURES + VAL_FEATURES))

# ==========================================
# 2. 視覺工具
# ==========================================
C_NAVY, C_BLUE, C_WHITE = '1F3864', '2E75B6', 'FFFFFF'

def _thin():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(cell, val, bg=C_BLUE, size=10):
    cell.value     = val
    cell.font      = Font(bold=True, color=C_WHITE, size=size)
    cell.fill      = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = _thin()

def tier_style(wr):
    if wr >= 0.65:   return PatternFill('solid',fgColor='C6EFCE'), Font(color='006100',bold=True), 'HIGH', '強'
    elif wr >= 0.55: return PatternFill('solid',fgColor='DDEBF7'), Font(color='1F497D',bold=True), 'MID',  '中'
    else:            return PatternFill('solid',fgColor='FFC7CE'), Font(color='9C0006',bold=True), 'LOW',  '弱'

def auto_width(ws, min_w=12, max_w=55):
    for col in ws.columns:
        best = min_w
        for cell in col:
            try:
                if cell.value:
                    best = min(max_w, max(best, len(str(cell.value)) + 4))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = best

# ==========================================
# 3. 宏觀過濾器
# ==========================================
def apply_macro_filter(pred_value, row_dict):
    vix    = row_dict.get('VIX指數',   20)
    sp_ret = row_dict.get('SP500_ret', 0)
    if vix > 25 or sp_ret < -0.05:
        return pred_value * 0.5 if pred_value > 0 else pred_value * 1.2
    return pred_value

# ==========================================
# 4. 特徵工程
# ==========================================
def build_features(df: pd.DataFrame, code: str = '') -> pd.DataFrame:
    d = df.copy()
    d['日期'] = pd.to_datetime(d['日期'])
    d = d.sort_values('日期').reset_index(drop=True)

    close = d['收盤價'].replace(0, np.nan)
    vol   = d['成交量'].replace(0, np.nan)

    d['MA5']    = close.rolling(5).mean().round(4)
    d['MA20']   = close.rolling(20).mean().round(4)
    d['ret5']   = close.pct_change(5).round(4)
    d['BIAS20'] = ((close - d['MA20']) / d['MA20'].replace(0, np.nan)).round(4)

    delta = close.diff()
    gain  = delta.clip(lower=0).rolling(14).mean()
    loss  = (-delta.clip(upper=0)).rolling(14).mean()
    d['RSI14'] = (100 - 100 / (1 + gain / loss.replace(0, np.nan))).round(4)

    if 'ATR' in d.columns and (d['ATR'] != 0).any():
        d['ATR14'] = d['ATR']
    elif '最高價' in d.columns and '最低價' in d.columns:
        high = d['最高價'].replace(0, np.nan)
        low  = d['最低價'].replace(0, np.nan)
        tr   = pd.concat([high - low,
                          (high - close.shift(1)).abs(),
                          (low  - close.shift(1)).abs()], axis=1).max(axis=1)
        d['ATR14'] = tr.rolling(14).mean().round(4)
    else:
        d['ATR14'] = close.diff().abs().rolling(14).mean().round(4)

    bb_mid = close.rolling(20).mean()
    bb_std = close.rolling(20).std()
    d['BB_width'] = ((bb_mid + 2*bb_std - (bb_mid - 2*bb_std))
                     / bb_mid.replace(0, np.nan)).round(4)
    d['量能比']  = (vol / vol.rolling(20).mean().replace(0, np.nan)).round(4)

    ema12 = close.ewm(span=12, adjust=False).mean()
    ema26 = close.ewm(span=26, adjust=False).mean()
    macd  = ema12 - ema26
    d['MACD_hist'] = (macd - macd.ewm(span=9, adjust=False).mean()).round(4)

    for col in ALL_FEATURES_SET:
        if col not in d.columns:
            d[col] = 0.0

    return d.ffill().bfill().dropna(subset=['收盤價', 'MA5', 'MA20'])

# ==========================================
# 5. Optuna 訓練
# ==========================================
def train_optuna_model(Xs: np.ndarray, ys: np.ndarray):
    if len(ys) < 25:
        return None

    def objective(trial):
        param = {
            'n_estimators':  trial.suggest_int('n_estimators',  50, 200),
            'max_depth':     trial.suggest_int('max_depth',       2,   5),
            'learning_rate': trial.suggest_float('learning_rate', 0.01, 0.15),
            'subsample':     trial.suggest_float('subsample',     0.6,  1.0),
            'random_state': 42, 'verbosity': 0,
        }
        split = int(len(Xs) * 0.8)
        reg   = XGBRegressor(**param).fit(Xs[:split], ys[:split])
        return float(np.mean((reg.predict(Xs[split:]) - ys[split:]) ** 2))

    study = optuna.create_study(direction='minimize')
    with tqdm(total=10, desc="     Optuna", unit="trial", ncols=72,
              bar_format="{desc}: {bar} {n}/{total} [{elapsed}<{remaining}]",
              leave=False) as pbar:
        def cb(study, trial): pbar.update(1)
        study.optimize(objective, n_trials=10, callbacks=[cb])

    model = Pipeline([
        ('s', StandardScaler()),
        ('m', XGBRegressor(**study.best_params, random_state=42, verbosity=0))
    ])
    model.fit(Xs, ys)
    return model

# ==========================================
# 6. 跨 ETF 訓練 + 動態信賴區間
# ==========================================
def train_and_predict(target_feat, others, horizon_days, feat_cols):
    Xs, ys = [], []
    for feat_df in others:
        tmp = feat_df.copy()
        tmp['target'] = tmp['收盤價'].pct_change(horizon_days).shift(-horizon_days)
        fc    = [f for f in feat_cols if f in tmp.columns]
        valid = tmp[fc + ['target']].dropna()
        valid = valid[valid['target'].abs() <= 0.6]
        if not valid.empty:
            Xs.append(valid[fc].values)
            ys.append(valid['target'].values)

    if not Xs:
        return 0.0, -0.10, 0.10, 0

    X_all = np.vstack(Xs)
    y_all = np.concatenate(ys)

    stds       = X_all.std(axis=0)
    valid_mask = stds > 0
    fc_valid   = [f for f, v in zip(feat_cols, valid_mask) if v]
    X_all      = X_all[:, valid_mask]

    model = train_optuna_model(X_all, y_all)
    if model is None:
        return 0.0, -0.10, 0.10, 0

    residuals = y_all - model.predict(X_all)
    lo_offset = np.percentile(residuals, 10)
    hi_offset = np.percentile(residuals, 90)

    target_fc  = [f for f in fc_valid if f in target_feat.columns]
    latest_val = target_feat[target_fc].iloc[-1:].values
    center     = float(model.predict(latest_val)[0])
    center     = apply_macro_filter(center, target_feat.iloc[-1].to_dict())

    return round(center, 4), round(center + lo_offset, 4), round(center + hi_offset, 4), len(fc_valid)

# ==========================================
# 7. 主流程
# ==========================================
def main():
    from pathlib import Path
    DATA_DIR = os.environ.get("ETF_DATA_DIR", str(Path.home() / "etf_data"))
    os.makedirs(DATA_DIR, exist_ok=True)

    _script_dir = os.path.dirname(os.path.abspath(__file__))
    def _find_file(filename):
        local = os.path.join(_script_dir, filename)
        return local if os.path.exists(local) else os.path.join(DATA_DIR, filename)

    DB_FILE  = _find_file('ETF歷史分析資料庫.xlsx')
    OUT_FILE = _find_file('ETF預測報告_v2.xlsx')

    if not os.path.exists(DB_FILE):
        print(f"❌ 找不到 {DB_FILE}，請先執行 etf_analysis.py")
        return

    # ── 讀取回測產出的預測力 ──────────────────────────────────────
    power_data = {}
    power_path = _find_file('etf_model_power.json')
    if os.path.exists(power_path):
        with open(power_path, encoding='utf-8') as f:
            power_data = json.load(f)
        print(f"📊 已載入 etf_model_power.json（{len(power_data)} 支）")
    else:
        print("ℹ️  未找到 etf_model_power.json，預測力顯示為「尚未回測」（建議先執行回測）")

    # ── 載入資料庫 ────────────────────────────────────────────────
    print(f"\n📂 載入：{DB_FILE}")
    xl = pd.ExcelFile(DB_FILE)
    all_dfs = {}
    for sheet in xl.sheet_names:
        code = sheet.split()[0].strip()
        try:
            df_raw = pd.read_excel(DB_FILE, sheet_name=sheet)
            all_dfs[code] = build_features(df_raw, code)
            print(f"   ✅ {sheet}: {len(all_dfs[code])} 筆")
        except Exception as e:
            print(f"   ❌ {sheet}: {e}")

    # ── 逐 ETF 預測 ───────────────────────────────────────────────
    results = []
    etf_bar = tqdm(all_dfs.items(), total=len(all_dfs),
                   desc="ETF 預測進度", unit="ETF", ncols=72,
                   bar_format="{desc}: {bar} {n}/{total}  [{elapsed}<{remaining}]")
    for idx, (code, df_feat) in enumerate(etf_bar, 1):
        name      = ETF_NAMES.get(code, code)
        category  = ETF_CATEGORY.get(code, '市值型')
        feat_cols = CATEGORY_FEATURES.get(category, DEFAULT_FEATURES)
        feat_cols = [f for f in feat_cols if f in df_feat.columns
                     and (df_feat[f] != 0).any()]
        others    = [d for c, d in all_dfs.items() if c != code]

        etf_bar.set_description(f"🔮 [{idx}/{len(all_dfs)}] {code} {name}")
        print(f"\n{'='*55}")
        print(f"🔮 [{idx}/{len(all_dfs)}] {code} {name}  [{category}, {len(feat_cols)}特徵]")
        print(f"{'='*55}")

        pred = {}
        feat_count = 0
        for hz_label, hz_days in tqdm([('3M', 63), ('6M', 126), ('12M', 252)],
                                         desc="  週期", unit="hz", ncols=72,
                                         bar_format="{desc}: {bar} {n}/{total}",
                                         leave=False):
            try:
                center, lo, hi, fc = train_and_predict(
                    df_feat, others, hz_days, feat_cols)
                pred[hz_label] = (center, lo, hi)
                feat_count = fc
                sig = '📈 偏多' if center > 0.05 else ('📉 偏空' if center < -0.05 else '➡️ 中性')
                print(f"   {hz_label}: {center:+.2%} [{lo:+.2%} ~ {hi:+.2%}]  {sig}")
            except Exception as e:
                print(f"   ⚠️ {hz_label} 失敗：{e}")
                pred[hz_label] = (0.0, -0.10, 0.10)

        pdata = power_data.get(code, {})
        results.append({
            'code':   code,
            'name':   name,
            'category': category,
            'price':  round(float(df_feat['收盤價'].iloc[-1]), 2),
            'date':   str(df_feat['日期'].iloc[-1].date()),
            'pred':   pred,
            'fc':     feat_count,
            'wr':     pdata.get('wr_12m', 0.0),
            'tier':   pdata.get('tier',  'LOW'),
            'stars':  pdata.get('stars', ''),
        })

    results.sort(key=lambda x: -x['wr'])

    # ── 寫 Excel ─────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = '📊 ETF預測總覽'

    ws.merge_cells('A1:S1')
    _hdr(ws['A1'],
         f'ETF AI 預測報告 v2.1　基準日：{datetime.now().strftime("%Y-%m-%d")}　預測力來自回測結果',
         bg=C_NAVY, size=13)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 36

    for ci, h in enumerate(['代號與名稱','ETF類型','預測力（回測）','12M方向勝率',
                             '現價','日期',
                             '3M預測中心','3M區間(Lo)','3M區間(Hi)','3M訊號',
                             '6M預測中心','6M區間(Lo)','6M區間(Hi)','6M訊號',
                             '12M預測中心','12M區間(Lo)','12M區間(Hi)','12M訊號',
                             '特徵數'], 1):
        _hdr(ws.cell(row=2, column=ci), h)

    hz_bg     = {'3M': 'E3F2FD', '6M': 'F3E5F5', '12M': 'E8F5E9'}
    sig_label = lambda v: '📈 偏多' if v > 0.05 else ('📉 偏空' if v < -0.05 else '➡️ 中性')
    sig_color = lambda v: '006100' if v > 0.05 else ('9C0006' if v < -0.05 else '7B5200')

    for ri, rec in enumerate(results, 3):
        tf, tf_font, _, tier_label = tier_style(rec['wr'])

        ws.cell(row=ri, column=1,
                value=f"{rec['code']} {rec['name']}").font = Font(bold=True)
        ws.cell(row=ri, column=2, value=rec['category'])

        stars_str = rec['stars'] or '尚未回測'
        label_str = f"{stars_str}  {tier_label}" if rec['stars'] else '尚未回測'
        pc = ws.cell(row=ri, column=3, value=label_str)
        if rec['stars']:
            pc.fill, pc.font = tf, tf_font
        pc.alignment = Alignment(horizontal='center')

        wc = ws.cell(row=ri, column=4,
                     value=f"{rec['wr']:.1%}" if rec['wr'] > 0 else '—')
        if rec['wr'] > 0:
            wc.fill, wc.font = tf, tf_font
        wc.alignment = Alignment(horizontal='center')

        ws.cell(row=ri, column=5, value=rec['price'])
        ws.cell(row=ri, column=6, value=rec['date'])

        col = 7
        for hz in ['3M', '6M', '12M']:
            center, lo, hi = rec['pred'][hz]
            bg = hz_bg[hz]
            for val in [center, lo, hi]:
                c = ws.cell(row=ri, column=col, value=val)
                c.number_format = '0.00%'
                c.fill   = PatternFill('solid', fgColor=bg)
                c.alignment = Alignment(horizontal='center')
                col += 1
            sc = ws.cell(row=ri, column=col, value=sig_label(center))
            sc.font      = Font(bold=True, color=sig_color(center))
            sc.fill      = PatternFill('solid', fgColor=bg)
            sc.alignment = Alignment(horizontal='center')
            col += 1

        ws.cell(row=ri, column=col, value=rec['fc'])

    auto_width(ws)
    wb.save(OUT_FILE)
    print(f"\n✨ 預測完成！報告：{OUT_FILE}")

    # ── 上傳 Drive ────────────────────────────────────────────────
    try:
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        from drive_sync import upload_file_path
        upload_file_path(os.path.abspath(OUT_FILE), "ETF預測報告_v2.xlsx")
    except ImportError:
        print("ℹ️  找不到 drive_sync.py，跳過上傳")
    except Exception as e:
        print(f"⚠️  Drive 上傳失敗（本地檔案完整）：{e}")


if __name__ == "__main__":
    main()
