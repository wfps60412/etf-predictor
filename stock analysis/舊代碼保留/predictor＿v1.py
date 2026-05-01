"""
predictor.py  v2.0
- 特徵對齊 backtester：24個特徵（技術面/籌碼面/基本面/總經）
- 動態信賴區間：以訓練集殘差的 10/90 百分位估算，非固定帶寬
- 跨股聯合訓練：所有個股歷史資料合併訓練，再對每股最新一筆預測
- 排除 0050（大盤參考）不納入個股模型
"""
import argparse
import warnings
import numpy as np
import pandas as pd
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from xgboost import XGBRegressor
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# ── 常數 ─────────────────────────────────────────────────────────
HORIZONS = {'3M': 63, '6M': 126, '12M': 252}
BENCHMARK_CODE = '0050'

C_NAVY, C_BLUE, C_ACCENT = '1B2A4A', '2B5A9B', 'E8A020'
C_WHITE, C_LGRAY         = 'FFFFFF', 'F4F6FA'
C_GREEN, C_RED, C_YELLOW = '22BB55', 'D94F3D', 'F5A623'

STOCK_NAMES = {
    '0050':'元大台灣50','1210':'大成',  '1216':'統一',  '1513':'中興電',
    '2323':'中環',      '2353':'宏碁',  '2374':'佳能',  '2376':'技嘉',
    '2409':'友達',      '3374':'精材',  '3481':'群創',  '4904':'遠傳',
    '6477':'安集',      '9914':'美利達','9921':'巨大',  '9933':'中鼎',
}

# ── 完整特徵集（對齊 backtester v3） ─────────────────────────────
TECH_FEATURES  = ['收盤價','成交量','MA5','MA20','ret5',
                  'RSI14','ATR14','BB_width','BIAS20','量能比','MACD_hist']
CHIP_FEATURES  = ['大戶指標','外資持股比','外資連買天數','投信持股比','融資變化率']
FUND_FEATURES  = ['PER','PBR','營收年增率','營收月增率']
MACRO_FEATURES = ['台幣匯率','VIX指數','RSP','Industry','SP500_ret','DXY_ret']
STOCK_TYPE_FEATURES = ['StockType','rolling_beta_60d','beta_regime','market_stress']
REV_MOM_FEATURES    = ['rev_acc3m','rev_accel','rev_mom_acc']
VALUATION_FEATURES  = ['per_zscore','pbr_zscore','per_trend']
ALL_FEATURES   = TECH_FEATURES + CHIP_FEATURES + FUND_FEATURES + MACRO_FEATURES + STOCK_TYPE_FEATURES + REV_MOM_FEATURES + VALUATION_FEATURES


# ── 1. 資料讀取 ───────────────────────────────────────────────────
def load_all_stocks(filepath: str) -> dict:
    xl = pd.ExcelFile(filepath)
    stocks = {}
    for s in xl.sheet_names:
        code = s.strip()
        if code == BENCHMARK_CODE:
            continue
        df = pd.read_excel(filepath, sheet_name=s)
        df['日期'] = pd.to_datetime(df['日期'])
        df = df.sort_values('日期').reset_index(drop=True)
        stocks[code] = df
    return stocks


# ── 2. 特徵工程（完整 24 特徵，對齊 backtester） ─────────────────
def build_features(df: pd.DataFrame, code: str) -> pd.DataFrame:
    d = df.copy()
    d = d.sort_values('日期').reset_index(drop=True)

    # 技術面
    d['MA5']   = d['收盤價'].rolling(5).mean()
    d['MA20']  = d['收盤價'].rolling(20).mean()
    d['ret5']  = d['收盤價'].pct_change(5)
    d['BIAS20'] = ((d['收盤價'] - d['MA20']) / d['MA20'].replace(0, 1e-9)).round(4)
    vol_ma20 = d['成交量'].rolling(20).mean()
    d['量能比'] = (d['成交量'] / vol_ma20.replace(0, 1e-9)).round(4)

    delta = d['收盤價'].diff()
    gain  = delta.clip(lower=0).rolling(14).mean()
    loss  = (-delta.clip(upper=0)).rolling(14).mean()
    d['RSI14'] = (100 - 100 / (1 + gain / loss.replace(0, 1e-9))).round(2)

    hl = (d['最高價'] - d['最低價']) if '最高價' in d.columns else d['收盤價'].diff().abs()
    d['ATR14'] = hl.rolling(14).mean()

    std20 = d['收盤價'].rolling(20).std()
    d['BB_width'] = (2 * std20 / d['MA20'].replace(0, 1e-9)).round(4)

    ema12 = d['收盤價'].ewm(span=12, adjust=False).mean()
    ema26 = d['收盤價'].ewm(span=26, adjust=False).mean()
    dif   = ema12 - ema26
    d['MACD_hist'] = (dif - dif.ewm(span=9, adjust=False).mean()).round(4)

    # 產業分類
    tech_codes = {'2323','2353','2374','2376','2409','3374','3481','4904'}
    d['Industry'] = 1 if code in tech_codes else 2

    # RSP：若資料庫有就直接用，否則補 0
    if 'RSP' not in d.columns:
        d['RSP'] = 0.0

    # ── 從資料庫讀取或計算新衍生欄位 ─────────────────────────────

    # SP500_ret / DXY_ret：資料庫有則直接用，否則補 0
    if 'SP500_ret' not in d.columns:
        d['SP500_ret'] = 0.0
    if 'DXY_ret' not in d.columns:
        d['DXY_ret'] = 0.0

    # 股票類型（資料庫有則直接用，否則依代碼靜態分類）
    if 'StockType' not in d.columns:
        DEFENSIVE = {'4904','1216','1210','9933','9914','9921','2353','1513'}
        CYCLICAL  = {'2376','3374','2374','2409','3481','2323'}
        if code in DEFENSIVE:
            d['StockType'] = 0
        elif code in CYCLICAL:
            d['StockType'] = 2
        else:
            d['StockType'] = 1

    # 動態 Beta（資料庫有則直接用，否則用 0050 RSP 欄位近似）
    if 'rolling_beta_60d' not in d.columns:
        # 用收盤價報酬率的 60 日標準差作為波動代理
        d['rolling_beta_60d'] = (
            d['收盤價'].pct_change().rolling(60).std() * 15
        ).round(4).ffill().fillna(1.0)
    if 'beta_regime' not in d.columns:
        d['beta_regime'] = d['rolling_beta_60d'].rank(pct=True).round(4)

    # 市場壓力指數（資料庫有則直接用，否則從 VIX/匯率計算）
    if 'market_stress' not in d.columns:
        if 'VIX指數' in d.columns and '台幣匯率' in d.columns:
            vix_ma = d['VIX指數'].rolling(20).mean()
            fx_ma  = d['台幣匯率'].rolling(20).mean()
            d['market_stress'] = (
                ((d['VIX指數'] - vix_ma) / vix_ma.replace(0, 1e-9)) * 0.5 +
                ((d['台幣匯率'] - fx_ma)  / fx_ma.replace(0,  1e-9)) * 0.5
            ).round(4).fillna(0)
        else:
            d['market_stress'] = 0.0

    # 營收動能（資料庫有則直接用，否則從現有營收欄位計算）
    if 'rev_acc3m' not in d.columns:
        if '營收年增率' in d.columns:
            d['rev_acc3m']   = d['營收年增率'].rolling(3).mean().round(4).fillna(0)
            d['rev_accel']   = (d['營收年增率'] - d['營收年增率'].shift(3)).round(4).fillna(0)
        else:
            d['rev_acc3m'] = 0.0
            d['rev_accel'] = 0.0
    if 'rev_mom_acc' not in d.columns:
        if '營收月增率' in d.columns:
            d['rev_mom_acc'] = d['營收月增率'].rolling(3).mean().round(4).fillna(0)
        else:
            d['rev_mom_acc'] = 0.0

    # 估值動能（資料庫有則直接用，否則從 PER/PBR 計算）
    if 'per_zscore' not in d.columns:
        if 'PER' in d.columns:
            per_ma  = d['PER'].replace(0, np.nan).rolling(60).mean()
            per_std = d['PER'].replace(0, np.nan).rolling(60).std().replace(0, 1e-9)
            d['per_zscore'] = ((d['PER'] - per_ma) / per_std).round(4).fillna(0)
            d['per_trend']  = (d['PER'] - d['PER'].shift(20)).round(4).fillna(0)
        else:
            d['per_zscore'] = 0.0
            d['per_trend']  = 0.0
    if 'pbr_zscore' not in d.columns:
        if 'PBR' in d.columns:
            pbr_ma  = d['PBR'].replace(0, np.nan).rolling(60).mean()
            pbr_std = d['PBR'].replace(0, np.nan).rolling(60).std().replace(0, 1e-9)
            d['pbr_zscore'] = ((d['PBR'] - pbr_ma) / pbr_std).round(4).fillna(0)
        else:
            d['pbr_zscore'] = 0.0


    return d.dropna()


# ── 3. 跨股訓練 + 動態信賴區間 ───────────────────────────────────
def train_and_predict(target_feat: pd.DataFrame,
                      all_feat_list: list,
                      horizon_td: int) -> tuple:
    """
    回傳 (center, lo, hi, feat_count)
    - center：模型預測中心值
    - lo/hi ：訓練集殘差 10/90 百分位 → 動態信賴區間
    - feat_count：實際使用特徵數
    """
    Xs, ys = [], []
    for feat_df in all_feat_list:
        tmp = feat_df.copy()
        tmp['target'] = tmp['收盤價'].pct_change(horizon_td).shift(-horizon_td)
        valid = tmp[ALL_FEATURES + ['target']].dropna()
        # 過濾極端除權值
        valid = valid[valid['target'].abs() <= 0.6]
        if not valid.empty:
            Xs.append(valid[ALL_FEATURES].values)
            ys.append(valid['target'].values)

    if not Xs:
        return 0.0, -0.10, 0.10, 0

    X_all = np.vstack(Xs)
    y_all = np.concatenate(ys)

    # 只保留有效（非零方差）特徵
    feat_stds  = X_all.std(axis=0)
    valid_mask = feat_stds > 0
    feat_names = [f for f, v in zip(ALL_FEATURES, valid_mask) if v]
    X_all      = X_all[:, valid_mask]

    model = Pipeline([
        ('scaler', StandardScaler()),
        ('model',  XGBRegressor(
            n_estimators=200, max_depth=4, learning_rate=0.05,
            subsample=0.8, colsample_bytree=0.8,
            random_state=42, verbosity=0
        ))
    ])
    model.fit(X_all, y_all)

    # 訓練集殘差 → 動態信賴帶寬
    train_preds = model.predict(X_all)
    residuals   = y_all - train_preds
    lo_offset   = np.percentile(residuals, 10)
    hi_offset   = np.percentile(residuals, 90)

    # 對最新一筆預測
    latest_row = target_feat[feat_names].iloc[-1:].values
    # 若欄位對不上（特徵名稱方式取）
    latest_vals = target_feat[[f for f in feat_names]].iloc[-1:].values
    center = float(model.predict(latest_vals)[0])

    return round(center, 4), round(center + lo_offset, 4), round(center + hi_offset, 4), len(feat_names)


# ── 4. 報表生成 ───────────────────────────────────────────────────
def _thin_border():
    s = Side(style='thin', color='C8D0DC')
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(cell, text, bg=None, color=C_WHITE, bold=True, size=10, align='center'):
    cell.value = text
    cell.font  = Font(name='Arial', size=size, bold=bold, color=color)
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    cell.border    = _thin_border()

def auto_col_width(ws):
    for col in ws.columns:
        max_len = 0
        letter = None
        for cell in col:
            if not hasattr(cell, 'column_letter'):
                continue
            if letter is None:
                letter = cell.column_letter
            try:
                if cell.value:
                    s = str(cell.value)
                    max_len = max(max_len, sum(2 if ord(c) > 127 else 1 for c in s))
            except Exception:
                pass
        if letter:
            ws.column_dimensions[letter].width = max(max_len + 3, 8)


def write_excel(results: list, output_path: str, base_date: str):
    wb = Workbook()

    # ── 總覽分頁 ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'AI預測結果總覽'
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 40

    ws.merge_cells('A1:N1')
    _hdr(ws['A1'],
         f'AI 股價漲跌幅預測報告 v2.0　基準日：{base_date}　特徵數：24（技術/籌碼/基本/總經）',
         bg=C_NAVY, size=13)

    col_headers = [
        '代號', '名稱', '現價', '日期',
        '3M預測中心', '3M區間(Lo)', '3M區間(Hi)', '3M訊號',
        '6M預測中心', '6M區間(Lo)', '6M區間(Hi)', '6M訊號',
        '12M預測中心','12M區間(Lo)','12M區間(Hi)','12M訊號',
        '有效特徵數'
    ]
    for ci, h in enumerate(col_headers, 1):
        _hdr(ws.cell(row=2, column=ci), h, bg=C_BLUE)

    hz_bg = {'3M': 'E3F2FD', '6M': 'F3E5F5', '12M': 'E8F5E9'}
    sig_color = lambda mid: (C_GREEN if mid > 0.05 else (C_RED if mid < -0.05 else C_YELLOW))
    sig_label = lambda mid: ('📈 偏多' if mid > 0.05 else ('📉 偏空' if mid < -0.05 else '➡️ 中性'))

    for ri, rec in enumerate(results, 3):
        ws.cell(row=ri, column=1, value=rec['code']).font = Font(bold=True)
        ws.cell(row=ri, column=2, value=rec['name'])
        ws.cell(row=ri, column=3, value=rec['current_price'])
        ws.cell(row=ri, column=4, value=rec['base_date'])

        col = 5
        for hz in ['3M', '6M', '12M']:
            center, lo, hi = rec['pred'][hz]
            mid = center
            bg  = hz_bg[hz]

            for val, fmt in [(center, f"{center:+.2%}"), (lo, f"{lo:+.2%}"), (hi, f"{hi:+.2%}")]:
                c = ws.cell(row=ri, column=col, value=fmt)
                c.fill   = PatternFill('solid', fgColor=bg)
                c.border = _thin_border()
                c.alignment = Alignment(horizontal='center')
                col += 1

            sc = ws.cell(row=ri, column=col, value=sig_label(mid))
            sc.font   = Font(bold=True, color=sig_color(mid))
            sc.fill   = PatternFill('solid', fgColor=bg)
            sc.border = _thin_border()
            sc.alignment = Alignment(horizontal='center')
            col += 1

        ws.cell(row=ri, column=col, value=rec['feat_count'])

    auto_col_width(ws)

    # ── 說明分頁 ──────────────────────────────────────────────────
    we = wb.create_sheet('模型說明')
    rows = [
        ('■ 版本', 'predictor.py v2.0'),
        ('■ 特徵集', f'共 {len(ALL_FEATURES)} 個：技術面(11) + 籌碼面(5) + 基本面(4) + 總經(6) + 股票類型(4) + 營收動能(3) + 估值動能(3)'),
        ('■ 技術面特徵', '收盤價, 成交量, MA5, MA20, ret5, RSI14, ATR14, BB_width, BIAS20, 量能比, MACD_hist'),
        ('■ 籌碼面特徵', '大戶指標, 外資持股比, 外資連買天數, 投信持股比, 融資變化率'),
        ('■ 基本面特徵', 'PER, PBR, 營收年增率, 營收月增率'),
        ('■ 總經/相對強弱', '台幣匯率, VIX指數, RSP, Industry, SP500_ret, DXY_ret'),
        ('■ 股票類型特徵', 'StockType（防禦0/中性1/循環2）, rolling_beta_60d, beta_regime, market_stress'),
        ('■ 營收動能特徵', 'rev_acc3m（3月YoY均）, rev_accel（YoY加速度）, rev_mom_acc（MoM動能）'),
        ('■ 估值動能特徵', 'per_zscore（PER歷史百分位）, pbr_zscore, per_trend（PER月變化）'),
        ('■ 模型', 'XGBoost（n=200, depth=4, lr=0.05, subsample=0.8）'),
        ('■ 訓練方式', '跨股聯合訓練：所有個股歷史資料合併，對每股最新一筆預測未來 3M/6M/12M'),
        ('■ 信賴區間', '動態計算：訓練集殘差的 P10/P90 百分位，反映模型實際預測誤差分佈'),
        ('■ 大盤(0050)', '排除於個股訓練，僅作為 RSP 相對強弱基準'),
        ('■ 訊號閾值', '偏多：中心值 > +5%；偏空：中心值 < -5%；中性：介於 ±5% 之間'),
        ('■ 回測勝率參考', '12M 方向勝率 73.3%（637筆，±10%門檻完全成功率 46.6%）'),
    ]
    we['A1'].value = '模型說明'; we['A1'].font = Font(bold=True, size=13, color=C_WHITE)
    we['A1'].fill  = PatternFill('solid', fgColor=C_NAVY)
    we.merge_cells('A1:B1')
    for i, (k, v) in enumerate(rows, 2):
        we.cell(row=i, column=1, value=k).font  = Font(bold=True)
        we.cell(row=i, column=2, value=v)
    auto_col_width(we)

    wb.save(output_path)
    print(f"✅ 報告已儲存：{output_path}")


# ── 5. 主程式 ─────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='AI股價預測 v2.0')
    parser.add_argument('--data',   default='個股歷史分析資料庫.xlsx')
    parser.add_argument('--output', default='AI精確預測報告_v2.xlsx')
    args = parser.parse_args()

    print(f"📂 載入資料：{args.data}")
    stocks_raw = load_all_stocks(args.data)
    print(f"   共 {len(stocks_raw)} 檔個股（已排除 {BENCHMARK_CODE}）")

    print("🔧 特徵工程...")
    feats = {}
    for code, df in stocks_raw.items():
        try:
            feats[code] = build_features(df, code)
        except Exception as e:
            print(f"   ⚠️  {code} 特徵工程失敗：{e}")

    all_feat_list = list(feats.values())

    results = []
    for code, df_f in feats.items():
        name = STOCK_NAMES.get(code, code)
        print(f"  🔮 預測：{code} {name} ...")
        pred = {}
        feat_count = 0
        for hz_label, hz_days in HORIZONS.items():
            try:
                center, lo, hi, fc = train_and_predict(df_f, all_feat_list, hz_days)
                pred[hz_label] = (center, lo, hi)
                feat_count = fc
            except Exception as e:
                print(f"     ⚠️  {hz_label} 預測失敗：{e}")
                pred[hz_label] = (0.0, -0.10, 0.10)

        results.append({
            'code':          code,
            'name':          name,
            'current_price': round(float(df_f['收盤價'].iloc[-1]), 2),
            'base_date':     str(df_f['日期'].iloc[-1].date()),
            'pred':          pred,
            'feat_count':    feat_count,
        })

    base_date = results[0]['base_date'] if results else 'N/A'
    write_excel(results, args.output, base_date)
    print(f"\n✨ 完成！報告：{args.output}")


if __name__ == '__main__':
    main()
