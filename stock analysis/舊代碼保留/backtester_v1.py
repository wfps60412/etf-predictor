"""
backtester.py  v4.0
改動：導入 predictor v2 的跨股聯合訓練邏輯
- 【舊】每支股票用自己 80% 歷史單獨訓練
- 【新】先用「所有其他股票的完整歷史」跨股聯合訓練，
        再對該股票的測試集預測，模擬真實部署情境
- 精準門檻同步提升至 ±10%（對齊 12M 回測討論結論）
- 總覽新增「方向勝率」欄位，與完全成功率並列
"""
import argparse
import pandas as pd
import numpy as np
from xgboost import XGBRegressor
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import warnings
warnings.filterwarnings('ignore')

# ─── 股票名稱 ────────────────────────────────────────────────────
STOCK_NAMES = {
    '0050':'元大台灣50','1210':'大成',  '1216':'統一',  '1513':'中興電',
    '2323':'中環',      '2353':'宏碁',  '2374':'佳能',  '2376':'技嘉',
    '2409':'友達',      '3374':'精材',  '3481':'群創',  '4904':'遠傳',
    '6477':'安集',      '9914':'美利達','9921':'巨大',  '9933':'中鼎',
}
BENCHMARK_CODE = '0050'

# ─── 視覺樣式 ────────────────────────────────────────────────────
C_NAVY, C_BLUE, C_WHITE = '1B2A4A', '2B5A9B', 'FFFFFF'
FILL_SUCCESS = PatternFill('solid', fgColor='C6EFCE'); FONT_SUCCESS = Font(color='006100', bold=True)
FILL_TREND   = PatternFill('solid', fgColor='FFEB9C'); FONT_TREND   = Font(color='9C5700', bold=True)
FILL_FAILED  = PatternFill('solid', fgColor='FFC7CE'); FONT_FAILED  = Font(color='9C0006', bold=True)
FILL_RIGHT   = PatternFill('solid', fgColor='E2EFDA')
FILL_WRONG   = PatternFill('solid', fgColor='FCE4D6')
HZ_COLORS    = {'3M':'E3F2FD', '6M':'F3E5F5', '12M':'E8F5E9'}
FILL_BENCH   = PatternFill('solid', fgColor='FFF3CD')

# ─── 核心參數 ────────────────────────────────────────────────────
DB_FILE       = '個股歷史分析資料庫.xlsx'
OUTPUT_FILE   = 'AI回測報告_v4_跨股訓練.xlsx'
HORIZONS      = {'3M': 63, '6M': 126, '12M': 252}
MIN_TEST_ROWS = 30
DIST_THRESHOLD = 0.10   # ±10%（從 ±5% 放寬，對齊 12M 回測討論）

# ─── 完整特徵集（24個，對齊 predictor v2）────────────────────────
TECH_FEATURES  = ['收盤價','成交量','MA5','MA20','ret5',
                  'RSI14','ATR14','BB_width','BIAS20','量能比','MACD_hist']
CHIP_FEATURES  = ['大戶指標','外資持股比','外資連買天數','投信持股比','融資變化率']
FUND_FEATURES  = ['PER','PBR','營收年增率','營收月增率']
MACRO_FEATURES = ['台幣匯率','VIX指數','RSP','Industry','SP500_ret','DXY_ret']
STOCK_TYPE_FEATURES = ['StockType','rolling_beta_60d','beta_regime','market_stress']
REV_MOM_FEATURES    = ['rev_acc3m','rev_accel','rev_mom_acc']
VALUATION_FEATURES  = ['per_zscore','pbr_zscore','per_trend']
ALL_FEATURES   = TECH_FEATURES + CHIP_FEATURES + FUND_FEATURES + MACRO_FEATURES + STOCK_TYPE_FEATURES + REV_MOM_FEATURES + VALUATION_FEATURES


# ─── 欄寬自動調整（相容合併儲存格）─────────────────────────────
def auto_adjust_width(ws):
    for col in ws.columns:
        max_len, letter = 0, None
        for cell in col:
            if not hasattr(cell, 'column_letter'):
                continue
            if letter is None:
                letter = cell.column_letter
            try:
                if cell.value:
                    s = str(cell.value)
                    max_len = max(max_len, sum(2 if ord(c) > 127 else 1 for c in s))
            except Exception:
                pass
        if letter:
            ws.column_dimensions[letter].width = max_len + 4


# ─── 特徵工程（完整 24 特徵）────────────────────────────────────
def build_features(df: pd.DataFrame, code: str) -> pd.DataFrame:
    d = df.copy()
    d['日期'] = pd.to_datetime(d['日期'])
    d = d.sort_values('日期').reset_index(drop=True)

    d['MA5']   = d['收盤價'].rolling(5).mean()
    d['MA20']  = d['收盤價'].rolling(20).mean()
    d['ret5']  = d['收盤價'].pct_change(5)
    d['BIAS20'] = ((d['收盤價'] - d['MA20']) / d['MA20'].replace(0, 1e-9)).round(4)
    vol_ma20   = d['成交量'].rolling(20).mean()
    d['量能比'] = (d['成交量'] / vol_ma20.replace(0, 1e-9)).round(4)

    delta = d['收盤價'].diff()
    gain  = delta.clip(lower=0).rolling(14).mean()
    loss  = (-delta.clip(upper=0)).rolling(14).mean()
    d['RSI14'] = (100 - 100 / (1 + gain / loss.replace(0, 1e-9))).round(2)

    hl = (d['最高價'] - d['最低價']) if '最高價' in d.columns else d['收盤價'].diff().abs()
    d['ATR14'] = hl.rolling(14).mean()

    std20 = d['收盤價'].rolling(20).std()
    d['BB_width'] = (2 * std20 / d['MA20'].replace(0, 1e-9)).round(4)

    ema12 = d['收盤價'].ewm(span=12, adjust=False).mean()
    ema26 = d['收盤價'].ewm(span=26, adjust=False).mean()
    dif   = ema12 - ema26
    d['MACD_hist'] = (dif - dif.ewm(span=9, adjust=False).mean()).round(4)

    tech_codes = {'2323','2353','2374','2376','2409','3374','3481','4904'}
    d['Industry'] = 1 if code in tech_codes else 2

    if 'RSP' not in d.columns:
        d['RSP'] = 0.0

    # ── 從資料庫讀取或計算新衍生欄位 ─────────────────────────────

    # SP500_ret / DXY_ret：資料庫有則直接用，否則補 0
    if 'SP500_ret' not in d.columns:
        d['SP500_ret'] = 0.0
    if 'DXY_ret' not in d.columns:
        d['DXY_ret'] = 0.0

    # 股票類型（資料庫有則直接用，否則依代碼靜態分類）
    if 'StockType' not in d.columns:
        DEFENSIVE = {'4904','1216','1210','9933','9914','9921','2353','1513'}
        CYCLICAL  = {'2376','3374','2374','2409','3481','2323'}
        if code in DEFENSIVE:
            d['StockType'] = 0
        elif code in CYCLICAL:
            d['StockType'] = 2
        else:
            d['StockType'] = 1

    # 動態 Beta（資料庫有則直接用，否則用 0050 RSP 欄位近似）
    if 'rolling_beta_60d' not in d.columns:
        # 用收盤價報酬率的 60 日標準差作為波動代理
        d['rolling_beta_60d'] = (
            d['收盤價'].pct_change().rolling(60).std() * 15
        ).round(4).ffill().fillna(1.0)
    if 'beta_regime' not in d.columns:
        d['beta_regime'] = d['rolling_beta_60d'].rank(pct=True).round(4)

    # 市場壓力指數（資料庫有則直接用，否則從 VIX/匯率計算）
    if 'market_stress' not in d.columns:
        if 'VIX指數' in d.columns and '台幣匯率' in d.columns:
            vix_ma = d['VIX指數'].rolling(20).mean()
            fx_ma  = d['台幣匯率'].rolling(20).mean()
            d['market_stress'] = (
                ((d['VIX指數'] - vix_ma) / vix_ma.replace(0, 1e-9)) * 0.5 +
                ((d['台幣匯率'] - fx_ma)  / fx_ma.replace(0,  1e-9)) * 0.5
            ).round(4).fillna(0)
        else:
            d['market_stress'] = 0.0

    # 營收動能（資料庫有則直接用，否則從現有營收欄位計算）
    if 'rev_acc3m' not in d.columns:
        if '營收年增率' in d.columns:
            d['rev_acc3m']   = d['營收年增率'].rolling(3).mean().round(4).fillna(0)
            d['rev_accel']   = (d['營收年增率'] - d['營收年增率'].shift(3)).round(4).fillna(0)
        else:
            d['rev_acc3m'] = 0.0
            d['rev_accel'] = 0.0
    if 'rev_mom_acc' not in d.columns:
        if '營收月增率' in d.columns:
            d['rev_mom_acc'] = d['營收月增率'].rolling(3).mean().round(4).fillna(0)
        else:
            d['rev_mom_acc'] = 0.0

    # 估值動能（資料庫有則直接用，否則從 PER/PBR 計算）
    if 'per_zscore' not in d.columns:
        if 'PER' in d.columns:
            per_ma  = d['PER'].replace(0, np.nan).rolling(60).mean()
            per_std = d['PER'].replace(0, np.nan).rolling(60).std().replace(0, 1e-9)
            d['per_zscore'] = ((d['PER'] - per_ma) / per_std).round(4).fillna(0)
            d['per_trend']  = (d['PER'] - d['PER'].shift(20)).round(4).fillna(0)
        else:
            d['per_zscore'] = 0.0
            d['per_trend']  = 0.0
    if 'pbr_zscore' not in d.columns:
        if 'PBR' in d.columns:
            pbr_ma  = d['PBR'].replace(0, np.nan).rolling(60).mean()
            pbr_std = d['PBR'].replace(0, np.nan).rolling(60).std().replace(0, 1e-9)
            d['pbr_zscore'] = ((d['PBR'] - pbr_ma) / pbr_std).round(4).fillna(0)
        else:
            d['pbr_zscore'] = 0.0


    return d.dropna()


# ─── 跨股聯合訓練（predictor v2 邏輯）──────────────────────────
def build_cross_stock_model(other_feat_dfs: list, hz_days: int) -> tuple:
    """
    用「其他所有股票」的完整歷史聯合訓練，回傳 (model, feat_names)
    other_feat_dfs：排除目標股後的所有 DataFrame list
    """
    Xs, ys = [], []
    for feat_df in other_feat_dfs:
        tmp = feat_df.copy()
        tmp['target'] = tmp['收盤價'].pct_change(hz_days).shift(-hz_days)
        valid = tmp[ALL_FEATURES + ['target']].dropna()
        valid = valid[valid['target'].abs() <= 0.6]   # 過濾除權極端值
        if not valid.empty:
            Xs.append(valid[ALL_FEATURES].values)
            ys.append(valid['target'].values)

    if not Xs:
        return None, []

    X_all = np.vstack(Xs)
    y_all = np.concatenate(ys)

    # 只保留非零方差特徵
    stds       = X_all.std(axis=0)
    valid_mask = stds > 0
    feat_names = [f for f, v in zip(ALL_FEATURES, valid_mask) if v]
    X_all      = X_all[:, valid_mask]

    model = Pipeline([
        ('s', StandardScaler()),
        ('m', XGBRegressor(
            n_estimators=200, max_depth=4, learning_rate=0.05,
            subsample=0.8, colsample_bytree=0.8,
            random_state=42, verbosity=0
        ))
    ])
    model.fit(X_all, y_all)
    return model, feat_names


# ─── 主回測流程 ──────────────────────────────────────────────────
def run_backtest(step_size: int):
    print(f"🚀 啟動回測 v4（跨股聯合訓練，step={step_size}，精準門檻±{DIST_THRESHOLD:.0%}）...")

    try:
        xl = pd.ExcelFile(DB_FILE)
    except Exception as e:
        print(f"❌ 找不到 {DB_FILE}：{e}"); return

    # ── 第一輪：全部讀入並做特徵工程 ────────────────────────────
    print("\n📦 第一輪：載入所有股票並建立特徵...")
    all_feats = {}   # code -> df_feat
    df_bench  = None

    for sheet_name in xl.sheet_names:
        code = sheet_name.strip()
        if code == BENCHMARK_CODE:
            df_bench = pd.read_excel(DB_FILE, sheet_name=sheet_name)
            df_bench['日期'] = pd.to_datetime(df_bench['日期'])
            continue
        try:
            df_raw = pd.read_excel(DB_FILE, sheet_name=sheet_name)
            all_feats[code] = build_features(df_raw, code)
            print(f"   ✅ {code} {STOCK_NAMES.get(code,'')}: {len(all_feats[code])} 筆")
        except Exception as e:
            print(f"   ⚠️  {code} 特徵工程失敗：{e}")

    # ── 建立報告 ─────────────────────────────────────────────────
    wb     = Workbook()
    ws_sum = wb.active
    ws_sum.title = '回測準確率總覽'

    sum_headers = ['股票','綜合評分',
                   '3M完全成功(±10%)', '3M方向勝率',
                   '6M完全成功(±10%)', '6M方向勝率',
                   '12M完全成功(±10%)','12M方向勝率',
                   '加權樣本數','有效特徵數']
    for ci, h in enumerate(sum_headers, 1):
        c = ws_sum.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color=C_WHITE)
        c.fill = PatternFill('solid', fgColor=C_BLUE)
        c.alignment = Alignment(horizontal='center')

    results_summary = []
    global_stats = {h: {'perf': [], 'trend': [], 'n': []} for h in HORIZONS}

    # ── 0050 大盤分頁 ────────────────────────────────────────────
    if df_bench is not None:
        ws_bench = wb.create_sheet('大盤分析_0050')
        _write_benchmark_sheet(ws_bench, df_bench, f"{BENCHMARK_CODE} 元大台灣50")
        print(f"\n📈 大盤分析分頁已建立（{len(df_bench)} 筆）")

    # ── 第二輪：逐股回測（跨股訓練）────────────────────────────
    print("\n🔮 第二輪：跨股聯合訓練回測...")
    for target_code, df_feat in all_feats.items():
        display_name = f"{target_code} {STOCK_NAMES.get(target_code,'未知')}"
        print(f"\n📊 回測：{display_name}")

        # 排除目標股，用其他所有股票訓練
        other_dfs = [df for code, df in all_feats.items() if code != target_code]

        ws_det = wb.create_sheet(f"{target_code}_明細")
        for ci, h in enumerate(['預測日期','週期','預測日收盤','預測漲幅%',
                                 '驗證日期','驗證日收盤','實際漲幅%','方向','誤差','綜合判定'], 1):
            c = ws_det.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color=C_WHITE)
            c.fill = PatternFill('solid', fgColor=C_NAVY)
            c.alignment = Alignment(horizontal='center')

        row_res = {'name': display_name, 'acc_info': {}, 'trend_info': {},
                   'weighted_score': 0, 'total_weight': 0, 'feat_count': 0}
        det_idx = 2

        for hz_label, hz_days in HORIZONS.items():

            # 跨股建模（排除目標股）
            model, feat_names = build_cross_stock_model(other_dfs, hz_days)
            if model is None:
                print(f"   ⚠️  {hz_label}: 跨股訓練資料不足，跳過")
                continue

            row_res['feat_count'] = len(feat_names)

            # 目標股：準備測試集
            d_task = df_feat.copy()
            d_task['target_price'] = d_task['收盤價'].shift(-hz_days)
            d_task['target_date']  = d_task['日期'].shift(-hz_days)
            d_task['actual_ret']   = (
                (d_task['收盤價'].shift(-hz_days) - d_task['收盤價']) / d_task['收盤價']
            )
            valid = d_task.dropna(subset=['actual_ret'])
            extreme = valid['actual_ret'].abs() > 0.6
            if extreme.sum() > 0:
                print(f"   ⚠️  {hz_label}: 過濾 {extreme.sum()} 筆除權極端值")
            valid = valid[~extreme]

            if len(valid) < 10:
                print(f"   ⚠️  {hz_label}: 有效資料不足10筆，跳過")
                continue

            # 全部當測試集（模型從未見過目標股）
            actual_step  = min(step_size, max(1, len(valid) // MIN_TEST_ROWS))
            test_stepped = valid.iloc[::actual_step]

            # 確保特徵欄位都存在
            feat_ok = [f for f in feat_names if f in test_stepped.columns]
            missing_feat = set(feat_names) - set(feat_ok)
            if missing_feat:
                print(f"   ℹ️  {hz_label}: 目標股缺少特徵 {missing_feat}，以0補足")
                for mf in missing_feat:
                    test_stepped = test_stepped.copy()
                    test_stepped[mf] = 0.0
                feat_ok = feat_names

            preds = model.predict(test_stepped[feat_ok])
            trues = test_stepped['actual_ret'].values

            same_dir = (preds * trues > 0)
            dist_ok  = (np.abs(trues - preds) <= DIST_THRESHOLD)
            perf  = np.mean(same_dir & dist_ok)
            trend = np.mean(same_dir)
            n     = len(test_stepped)

            row_res['acc_info'][hz_label]   = f"{perf:.0%}"
            row_res['trend_info'][hz_label] = f"{trend:.0%}"
            hz_score = (np.sum(same_dir & dist_ok) * 1.0 + np.sum(same_dir & ~dist_ok) * 0.5) / n
            row_res['weighted_score'] += hz_score * n
            row_res['total_weight']   += n
            global_stats[hz_label]['perf'].append(perf)
            global_stats[hz_label]['trend'].append(trend)
            global_stats[hz_label]['n'].append(n)

            print(f"   {hz_label}: 方向勝率={trend:.1%}  完全成功(±10%)={perf:.1%}  樣本={n}  特徵={len(feat_ok)}")

            # ── 寫明細 ────────────────────────────────────────────
            start_row = det_idx
            for i in range(len(test_stepped)):
                row = test_stepped.iloc[i]
                ws_det.cell(row=det_idx, column=1, value=row['日期'].strftime('%Y-%m-%d'))
                ws_det.cell(row=det_idx, column=2, value=hz_label)
                ws_det.cell(row=det_idx, column=3, value=round(float(row['收盤價']), 2))
                ws_det.cell(row=det_idx, column=4, value=f"{preds[i]:+.2%}")
                ws_det.cell(row=det_idx, column=5, value=row['target_date'].strftime('%Y-%m-%d'))
                ws_det.cell(row=det_idx, column=6, value=round(float(row['target_price']), 2))
                ws_det.cell(row=det_idx, column=7, value=f"{trues[i]:+.2%}")

                cd = ws_det.cell(row=det_idx, column=8, value="正確" if same_dir[i] else "錯誤")
                cd.fill = FILL_RIGHT if same_dir[i] else FILL_WRONG
                ce = ws_det.cell(row=det_idx, column=9, value="精準" if dist_ok[i] else "落後")
                ce.fill = FILL_RIGHT if dist_ok[i] else FILL_WRONG

                if same_dir[i] and dist_ok[i]:
                    lbl, fill, font = "完全成功", FILL_SUCCESS, FONT_SUCCESS
                elif same_dir[i]:
                    lbl, fill, font = "趨勢正確", FILL_TREND, FONT_TREND
                else:
                    lbl, fill, font = "預測失敗", FILL_FAILED, FONT_FAILED
                cr = ws_det.cell(row=det_idx, column=10, value=lbl)
                cr.fill, cr.font = fill, font
                det_idx += 1

            if det_idx - 1 >= start_row:
                ws_det.merge_cells(start_row=start_row, start_column=2,
                                   end_row=det_idx-1, end_column=2)
                tc = ws_det.cell(row=start_row, column=2)
                tc.value = hz_label
                tc.fill  = PatternFill('solid', fgColor=HZ_COLORS[hz_label])
                tc.alignment = Alignment(horizontal='center', vertical='center')

        row_res['final_score'] = (row_res['weighted_score'] / row_res['total_weight']
                                  if row_res['total_weight'] > 0 else 0)
        results_summary.append(row_res)
        auto_adjust_width(ws_det)

    # ── 總覽填充 ──────────────────────────────────────────────────
    for ri, rec in enumerate(results_summary, 2):
        ws_sum.cell(row=ri, column=1, value=rec['name']).font = Font(bold=True)
        ws_sum.cell(row=ri, column=2, value=f"{rec['final_score']:.3f}").font = Font(bold=True, color='0000FF')
        col = 3
        for hz in ['3M', '6M', '12M']:
            ws_sum.cell(row=ri, column=col,   value=rec['acc_info'].get(hz,   '-'))
            ws_sum.cell(row=ri, column=col+1, value=rec['trend_info'].get(hz, '-'))
            col += 2
        ws_sum.cell(row=ri, column=9,  value=rec['total_weight'])
        ws_sum.cell(row=ri, column=10, value=rec['feat_count'])

    # 加權平均列
    curr_row = len(results_summary) + 3
    cl = ws_sum.cell(row=curr_row, column=1, value="模型加權平均 (完全成功±10% / 方向勝率)")
    cl.font = Font(bold=True, color=C_WHITE)
    cl.fill = PatternFill('solid', fgColor=C_NAVY)
    col = 3
    for hz in ['3M', '6M', '12M']:
        ns = global_stats[hz]['n']
        if not ns:
            col += 2; continue
        wp = np.average(global_stats[hz]['perf'],  weights=ns)
        wt = np.average(global_stats[hz]['trend'], weights=ns)
        ws_sum.cell(row=curr_row, column=col,   value=f"{wp:.1%}").font = Font(bold=True, color='FF0000')
        ws_sum.cell(row=curr_row, column=col+1, value=f"{wt:.1%}").font = Font(bold=True, color='FF0000')
        col += 2

    auto_adjust_width(ws_sum)
    wb.save(OUTPUT_FILE)
    print(f"\n✨ 完成！報告：{OUTPUT_FILE}")


def _write_benchmark_sheet(ws, df, display_name):
    ws.merge_cells('A1:F1')
    c = ws.cell(row=1, column=1, value=f"{display_name} — 大盤參考（不納入個股模型訓練）")
    c.font = Font(bold=True, color='9C5700'); c.fill = FILL_BENCH
    c.alignment = Alignment(horizontal='center')
    ws.cell(row=2, column=1, value='此分頁僅供觀察大盤趨勢，作為個股 RSP（相對強弱）的計算基準。')
    ws.cell(row=2, column=1).font = Font(italic=True, color='595959', size=10)

    cols = ['日期','開盤價','最高價','最低價','收盤價','漲跌價差','成交量','大戶指標','台幣匯率','VIX指數']
    for ci, h in enumerate(cols, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = Font(bold=True, color=C_WHITE)
        c.fill = PatternFill('solid', fgColor='8B6914')
        c.alignment = Alignment(horizontal='center')

    for ri, (_, row) in enumerate(df.iterrows(), 5):
        for ci, col in enumerate(cols, 1):
            if col in df.columns:
                val = row[col]
                if hasattr(val, 'strftime'):
                    val = val.strftime('%Y-%m-%d')
                elif isinstance(val, (float, np.floating)):
                    val = round(float(val), 4)
                ws.cell(row=ri, column=ci, value=val)
    auto_adjust_width(ws)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='AI回測 v4 - 跨股聯合訓練')
    parser.add_argument('--step', type=int, default=3,
                        help='取樣步長（預設3）')
    args = parser.parse_args()
    run_backtest(args.step)
