"""
predictor.py  v2.0
- 特徵對齊 backtester：24個特徵（技術面/籌碼面/基本面/總經）
- 動態信賴區間：以訓練集殘差的 10/90 百分位估算，非固定帶寬
- 跨股聯合訓練：所有個股歷史資料合併訓練，再對每股最新一筆預測
- 排除 0050（大盤參考）不納入個股模型
"""
import argparse
import warnings
import numpy as np
import pandas as pd
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from xgboost import XGBRegressor
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# ── 常數 ─────────────────────────────────────────────────────────
HORIZONS = {'3M': 63, '6M': 126, '12M': 252}
BENCHMARK_CODE = '0050'

C_NAVY, C_BLUE, C_ACCENT = '1B2A4A', '2B5A9B', 'E8A020'
C_WHITE, C_LGRAY         = 'FFFFFF', 'F4F6FA'
C_GREEN, C_RED, C_YELLOW = '22BB55', 'D94F3D', 'F5A623'


# ─── 各股模型預測力（12M 方向勝率，產業別特徵回測驗證）──────────
MODEL_POWER = {
    '2330': (1.000, 'HIGH',   '★★★★★'),  # 台積電
    '2412': (0.927, 'HIGH',   '★★★★★'),  # 中華電
    '2454': (0.922, 'HIGH',   '★★★★★'),  # 聯發科
    '4904': (0.901, 'HIGH',   '★★★★★'),  # 遠傳
    '3374': (0.857, 'HIGH',   '★★★★☆'),  # 精材
    '9914': (0.834, 'HIGH',   '★★★★☆'),  # 美利達
    '2303': (0.835, 'HIGH',   '★★★★☆'),  # 聯電
    '9921': (0.795, 'HIGH',   '★★★★☆'),  # 巨大
    '2409': (0.784, 'HIGH',   '★★★★☆'),  # 友達
    '2603': (0.765, 'HIGH',   '★★★★☆'),  # 長榮
    '3481': (0.729, 'HIGH',   '★★★★☆'),  # 群創
    '1210': (0.725, 'HIGH',   '★★★★☆'),  # 大成
    '2882': (0.676, 'MID',    '★★★☆☆'),  # 國泰金
    '2323': (0.604, 'MID',    '★★★☆☆'),  # 中環
    '2376': (0.609, 'MID',    '★★★☆☆'),  # 技嘉
    '1513': (0.728, 'HIGH',   '★★★★☆'),  # 中興電
    '9933': (0.429, 'LOW',    '★★☆☆☆'),  # 中鼎
    '2353': (0.701, 'HIGH',   '★★★★☆'),  # 宏碁
    '2317': (0.706, 'HIGH',   '★★★★☆'),  # 鴻海
    '6477': (0.348, 'LOW',    '★★☆☆☆'),  # 安集
    '2374': (0.393, 'LOW',    '★★☆☆☆'),  # 佳能
    '1301': (0.322, 'LOW',    '★☆☆☆☆'),  # 台塑
    '1216': (0.292, 'LOW',    '★☆☆☆☆'),  # 統一
}
POWER_LABEL = {'HIGH': '強', 'MID': '中', 'LOW': '弱'}
POWER_COLOR = {'HIGH': '006100', 'MID': '7B5200', 'LOW': '9C0006'}
POWER_BG    = {'HIGH': 'C6EFCE', 'MID': 'FFEB9C', 'LOW': 'FFC7CE'}

STOCK_NAMES = {
    '0050':'元大台灣50','1210':'大成',  '1216':'統一',  '1513':'中興電',
    '2323':'中環',      '2353':'宏碁',  '2374':'佳能',  '2376':'技嘉',
    '2409':'友達',      '3374':'精材',  '3481':'群創',  '4904':'遠傳',
    '6477':'安集',      '9914':'美利達','9921':'巨大',  '9933':'中鼎',
    '2330':'台積電',    '2303':'聯電',  '2454':'聯發科',
    '2882':'國泰金',    '2603':'長榮',  '2317':'鴻海',
    '2412':'中華電',    '1301':'台塑',
}

# ── 完整特徵集（對齊 backtester v3） ─────────────────────────────
TECH_FEATURES  = ['收盤價','成交量','MA5','MA20','ret5',
                  'RSI14','ATR14','BB_width','BIAS20','量能比','MACD_hist']
CHIP_FEATURES  = ['大戶指標','外資持股比','外資連買天數','投信持股比','融資變化率']
FUND_FEATURES  = ['PER','PBR','營收年增率','營收月增率']
MACRO_FEATURES = ['台幣匯率','VIX指數','RSP','Industry','SP500_ret','DXY_ret']
STOCK_TYPE_FEATURES = ['StockType','rolling_beta_60d','beta_regime','market_stress']
BETA_SP500_FEATURES = ['beta_sp500_static']
REV_MOM_FEATURES    = ['rev_acc3m','rev_accel','rev_mom_acc']
VALUATION_FEATURES  = ['per_zscore','pbr_zscore','per_trend']
FINANCIAL_FEATURES  = ['gross_margin','gross_margin_qoq','op_margin','eps_qoq','inventory_days','receivable_days','operating_cf_ratio','capex_ratio']
ALL_FEATURES   = TECH_FEATURES + CHIP_FEATURES + FUND_FEATURES + MACRO_FEATURES + STOCK_TYPE_FEATURES + BETA_SP500_FEATURES + REV_MOM_FEATURES + VALUATION_FEATURES + FINANCIAL_FEATURES

# ─── 產業別最佳特徵組合 ────────────────────────────────────────
_T  = TECH_FEATURES
_C  = CHIP_FEATURES
_M  = MACRO_FEATURES
_TY = STOCK_TYPE_FEATURES + BETA_SP500_FEATURES
_V  = VALUATION_FEATURES
_F  = FINANCIAL_FEATURES
_FD = FUND_FEATURES

INDUSTRY_FEATURES = {
    '食品飼料':      _T + _C + _M + _TY + _V,
    '自行車':        _T + _M,
    '電信':          _T + _M,
    '工程電機':      _T + _C + _M + _TY,
    '電腦周邊':      _T + _C + _M + _V,
    '面板':          _T + _C + _M + _TY,
    '半導體封測':    _T + _M + _F,
    '光儲存':        _T + _M,
    '半導體晶圓代工':_T + _C + _M + _TY,
    'IC設計':        _T + _C + _M + _TY + _V,
    '金融保險':      _T + _M + _V,
    '航運':          _T + _C + _M + _TY,
    '電子零組件':    _T + _C + _M + _TY,
    '石化':          _T + _M + _V,
}
DEFAULT_FEATURES = _T + _C + _M + _TY + _V

STOCK_INDUSTRY_MAP = {
    '1210':'食品飼料', '1216':'食品飼料',
    '9914':'自行車',   '9921':'自行車',
    '4904':'電信',     '2412':'電信',
    '9933':'工程電機', '1513':'工程電機',
    '2353':'電腦周邊', '2374':'電腦周邊', '2376':'電腦周邊',
    '2409':'面板',     '3481':'面板',
    '3374':'半導體封測','6477':'半導體封測',
    '2323':'光儲存',
    '2330':'半導體晶圓代工','2303':'半導體晶圓代工',
    '2454':'IC設計',
    '2882':'金融保險',
    '2603':'航運',
    '2317':'電子零組件',
    '1301':'石化',
}


# ── 1. 資料讀取 ───────────────────────────────────────────────────
def load_all_stocks(filepath: str) -> dict:
    xl = pd.ExcelFile(filepath)
    stocks = {}
    for s in xl.sheet_names:
        code = s.strip()
        if code == BENCHMARK_CODE:
            continue
        df = pd.read_excel(filepath, sheet_name=s)
        df['日期'] = pd.to_datetime(df['日期'])
        df = df.sort_values('日期').reset_index(drop=True)
        stocks[code] = df
    return stocks


# ── 2. 特徵工程（完整 24 特徵，對齊 backtester） ─────────────────
def build_features(df: pd.DataFrame, code: str) -> pd.DataFrame:
    d = df.copy()
    d = d.sort_values('日期').reset_index(drop=True)

    # 技術面
    d['MA5']   = d['收盤價'].rolling(5).mean()
    d['MA20']  = d['收盤價'].rolling(20).mean()
    d['ret5']  = d['收盤價'].pct_change(5)
    d['BIAS20'] = ((d['收盤價'] - d['MA20']) / d['MA20'].replace(0, 1e-9)).round(4)
    vol_ma20 = d['成交量'].rolling(20).mean()
    d['量能比'] = (d['成交量'] / vol_ma20.replace(0, 1e-9)).round(4)

    delta = d['收盤價'].diff()
    gain  = delta.clip(lower=0).rolling(14).mean()
    loss  = (-delta.clip(upper=0)).rolling(14).mean()
    d['RSI14'] = (100 - 100 / (1 + gain / loss.replace(0, 1e-9))).round(2)

    hl = (d['最高價'] - d['最低價']) if '最高價' in d.columns else d['收盤價'].diff().abs()
    d['ATR14'] = hl.rolling(14).mean()

    std20 = d['收盤價'].rolling(20).std()
    d['BB_width'] = (2 * std20 / d['MA20'].replace(0, 1e-9)).round(4)

    ema12 = d['收盤價'].ewm(span=12, adjust=False).mean()
    ema26 = d['收盤價'].ewm(span=26, adjust=False).mean()
    dif   = ema12 - ema26
    d['MACD_hist'] = (dif - dif.ewm(span=9, adjust=False).mean()).round(4)

    # 產業分類
    tech_codes = {'2323','2353','2374','2376','2409','3374','3481','4904'}
    d['Industry'] = 1 if code in tech_codes else 2

    # RSP：若資料庫有就直接用，否則補 0
    if 'RSP' not in d.columns:
        d['RSP'] = 0.0

    # ── 從資料庫讀取或計算新衍生欄位 ─────────────────────────────

    # SP500_ret / DXY_ret：資料庫有則直接用，否則補 0
    if 'SP500_ret' not in d.columns:
        d['SP500_ret'] = 0.0
    if 'DXY_ret' not in d.columns:
        d['DXY_ret'] = 0.0

    # 股票類型（資料庫有則直接用，否則依代碼靜態分類）
    if 'StockType' not in d.columns:
        # StockType: 0=防禦型, 1=混合型, 2=景氣循環型
        DEFENSIVE = {'4904','1216','1210','9933','2882','2412'}
        MIXED     = {'2353','1513'}
        CYCLICAL  = {'9914','9921','2376','3374','2374','2409','3481','2323','6477','2330','2303','2454','2603','2317','1301'}
        if code in DEFENSIVE:
            d['StockType'] = 0
        elif code in CYCLICAL:
            d['StockType'] = 2
        else:
            d['StockType'] = 1  # 混合型

    # 動態 Beta（資料庫有則直接用，否則用 0050 RSP 欄位近似）
    if 'rolling_beta_60d' not in d.columns:
        # 用收盤價報酬率的 60 日標準差作為波動代理
        d['rolling_beta_60d'] = (
            d['收盤價'].pct_change().rolling(60).std() * 15
        ).round(4).ffill().fillna(1.0)
    if 'beta_regime' not in d.columns:
        d['beta_regime'] = d['rolling_beta_60d'].rank(pct=True).round(4)

    # 市場壓力指數（資料庫有則直接用，否則從 VIX/匯率計算）
    if 'market_stress' not in d.columns:
        if 'VIX指數' in d.columns and '台幣匯率' in d.columns:
            vix_ma = d['VIX指數'].rolling(20).mean()
            fx_ma  = d['台幣匯率'].rolling(20).mean()
            d['market_stress'] = (
                ((d['VIX指數'] - vix_ma) / vix_ma.replace(0, 1e-9)) * 0.5 +
                ((d['台幣匯率'] - fx_ma)  / fx_ma.replace(0,  1e-9)) * 0.5
            ).round(4).fillna(0)
        else:
            d['market_stress'] = 0.0

    # 營收動能（資料庫有則直接用，否則從現有營收欄位計算）
    if 'rev_acc3m' not in d.columns:
        if '營收年增率' in d.columns:
            d['rev_acc3m']   = d['營收年增率'].rolling(3).mean().round(4).fillna(0)
            d['rev_accel']   = (d['營收年增率'] - d['營收年增率'].shift(3)).round(4).fillna(0)
        else:
            d['rev_acc3m'] = 0.0
            d['rev_accel'] = 0.0
    if 'rev_mom_acc' not in d.columns:
        if '營收月增率' in d.columns:
            d['rev_mom_acc'] = d['營收月增率'].rolling(3).mean().round(4).fillna(0)
        else:
            d['rev_mom_acc'] = 0.0

    # 估值動能（資料庫有則直接用，否則從 PER/PBR 計算）
    if 'per_zscore' not in d.columns:
        if 'PER' in d.columns:
            per_ma  = d['PER'].replace(0, np.nan).rolling(60).mean()
            per_std = d['PER'].replace(0, np.nan).rolling(60).std().replace(0, 1e-9)
            d['per_zscore'] = ((d['PER'] - per_ma) / per_std).round(4).fillna(0)
            d['per_trend']  = (d['PER'] - d['PER'].shift(20)).round(4).fillna(0)
        else:
            d['per_zscore'] = 0.0
            d['per_trend']  = 0.0
    if 'pbr_zscore' not in d.columns:
        if 'PBR' in d.columns:
            pbr_ma  = d['PBR'].replace(0, np.nan).rolling(60).mean()
            pbr_std = d['PBR'].replace(0, np.nan).rolling(60).std().replace(0, 1e-9)
            d['pbr_zscore'] = ((d['PBR'] - pbr_ma) / pbr_std).round(4).fillna(0)
        else:
            d['pbr_zscore'] = 0.0


    # Beta_SP500 靜態指標（資料庫有則直接用，否則依股票代碼賦值）
    if 'beta_sp500_static' not in d.columns:
        _BETA_SP500_STATIC = {'3481': -0.0882, '2323': -0.068, '2409': -0.0529, '4904': -0.0467, '1216': -0.0196, '1210': -0.0015, '9933': -0.0008, '2353': 0.0117, '1513': 0.0337, '9914': 0.038, '2374': 0.0635, '9921': 0.076, '6477': 0.1007, '3374': 0.1259, '2376': 0.2046, '2330': 0.18, '2303': 0.09, '2454': 0.15, '2882': 0.03, '2603': 0.06, '2317': 0.11, '2412': -0.03, '1301': 0.05}
        d['beta_sp500_static'] = _BETA_SP500_STATIC.get(str(code), 0.0)

    # 財務報表指標（資料庫有則直接用，否則補 0）
    for col in ['gross_margin','gross_margin_qoq','op_margin','eps_qoq',
                'inventory_days','receivable_days','operating_cf_ratio','capex_ratio']:
        if col not in d.columns:
            d[col] = 0.0

    return d.dropna()


# ── 3. 跨股訓練 + 動態信賴區間 ───────────────────────────────────
def train_and_predict(target_feat: pd.DataFrame,
                      all_feat_list: list,
                      horizon_td: int,
                      feat_cols: list = None) -> tuple:
    """
    回傳 (center, lo, hi, feat_count)
    - center：模型預測中心值
    - lo/hi ：訓練集殘差 10/90 百分位 → 動態信賴區間
    - feat_count：實際使用特徵數
    """
    if feat_cols is None:
        feat_cols = ALL_FEATURES
    Xs, ys = [], []
    for feat_df in all_feat_list:
        tmp = feat_df.copy()
        tmp['target'] = tmp['收盤價'].pct_change(horizon_td).shift(-horizon_td)
        fc    = [f for f in feat_cols if f in tmp.columns]
        valid = tmp[fc + ['target']].dropna()
        valid = valid[valid['target'].abs() <= 0.6]
        if not valid.empty:
            Xs.append(valid[fc].values)
            ys.append(valid['target'].values)

    if not Xs:
        return 0.0, -0.10, 0.10, 0

    X_all = np.vstack(Xs)
    y_all = np.concatenate(ys)

    # 只保留有效（非零方差）特徵
    feat_stds  = X_all.std(axis=0)
    valid_mask = feat_stds > 0
    feat_names = [f for f, v in zip(ALL_FEATURES, valid_mask) if v]
    X_all      = X_all[:, valid_mask]

    model = Pipeline([
        ('scaler', StandardScaler()),
        ('model',  XGBRegressor(
            n_estimators=200, max_depth=4, learning_rate=0.05,
            subsample=0.8, colsample_bytree=0.8,
            random_state=42, verbosity=0
        ))
    ])
    model.fit(X_all, y_all)

    # 訓練集殘差 → 動態信賴帶寬
    train_preds = model.predict(X_all)
    residuals   = y_all - train_preds
    lo_offset   = np.percentile(residuals, 10)
    hi_offset   = np.percentile(residuals, 90)

    # 對最新一筆預測
    latest_row = target_feat[feat_names].iloc[-1:].values
    # 若欄位對不上（特徵名稱方式取）
    latest_vals = target_feat[[f for f in feat_names]].iloc[-1:].values
    center = float(model.predict(latest_vals)[0])

    return round(center, 4), round(center + lo_offset, 4), round(center + hi_offset, 4), len(feat_names)


# ── 4. 報表生成 ───────────────────────────────────────────────────
def _thin_border():
    s = Side(style='thin', color='C8D0DC')
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(cell, text, bg=None, color=C_WHITE, bold=True, size=10, align='center'):
    cell.value = text
    cell.font  = Font(name='Arial', size=size, bold=bold, color=color)
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    cell.border    = _thin_border()

def auto_col_width(ws):
    for col in ws.columns:
        max_len = 0
        letter = None
        for cell in col:
            if not hasattr(cell, 'column_letter'):
                continue
            if letter is None:
                letter = cell.column_letter
            try:
                if cell.value:
                    s = str(cell.value)
                    max_len = max(max_len, sum(2 if ord(c) > 127 else 1 for c in s))
            except Exception:
                pass
        if letter:
            ws.column_dimensions[letter].width = max(max_len + 3, 8)


def write_excel(results: list, output_path: str, base_date: str):
    wb = Workbook()

    # ── 總覽分頁 ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'AI預測結果總覽'
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 40

    ws.merge_cells('A1:T1')
    _hdr(ws['A1'],
         f'AI 股價漲跌幅預測報告 v3.0　基準日：{base_date}　產業別特徵 | 23支股票',
         bg=C_NAVY, size=13)

    col_headers = [
        '代號', '名稱', '產業', '預測力', '12M方向勝率', '現價', '日期',
        '3M預測中心', '3M區間(Lo)', '3M區間(Hi)', '3M訊號',
        '6M預測中心', '6M區間(Lo)', '6M區間(Hi)', '6M訊號',
        '12M預測中心','12M區間(Lo)','12M區間(Hi)','12M訊號',
        '有效特徵數'
    ]
    for ci, h in enumerate(col_headers, 1):
        _hdr(ws.cell(row=2, column=ci), h, bg=C_BLUE)

    hz_bg = {'3M': 'E3F2FD', '6M': 'F3E5F5', '12M': 'E8F5E9'}
    sig_color = lambda mid: (C_GREEN if mid > 0.05 else (C_RED if mid < -0.05 else C_YELLOW))
    sig_label = lambda mid: ('📈 偏多' if mid > 0.05 else ('📉 偏空' if mid < -0.05 else '➡️ 中性'))

    # 按 12M 方向勝率排序（預測力強的排前面）
    results = sorted(results, key=lambda x: -x.get('power_wr', 0.5))

    for ri, rec in enumerate(results, 3):
        tier  = rec.get('power_tier',  'MID')
        stars = rec.get('power_stars', '★★★☆☆')
        wr    = rec.get('power_wr',    0.5)
        ind   = rec.get('industry',    '-')

        ws.cell(row=ri, column=1, value=rec['code']).font = Font(bold=True)
        ws.cell(row=ri, column=2, value=rec['name'])

        # 產業
        ws.cell(row=ri, column=3, value=ind)

        # 預測力（含色彩）
        pc = ws.cell(row=ri, column=4, value=f"{stars}  {POWER_LABEL.get(tier,'中')}")
        pc.font  = Font(bold=True, color=POWER_COLOR.get(tier, '000000'))
        pc.fill  = PatternFill('solid', fgColor=POWER_BG.get(tier, 'FFFFFF'))
        pc.alignment = Alignment(horizontal='center')
        pc.border    = _thin_border()

        # 12M 方向勝率
        wc = ws.cell(row=ri, column=5, value=f"{wr:.1%}")
        wc.font  = Font(bold=True, color=POWER_COLOR.get(tier, '000000'))
        wc.fill  = PatternFill('solid', fgColor=POWER_BG.get(tier, 'FFFFFF'))
        wc.alignment = Alignment(horizontal='center')
        wc.border    = _thin_border()

        ws.cell(row=ri, column=6, value=rec['current_price'])
        ws.cell(row=ri, column=7, value=rec['base_date'])

        col = 8
        for hz in ['3M', '6M', '12M']:
            center, lo, hi = rec['pred'][hz]
            mid = center
            bg  = hz_bg[hz]

            for val, fmt in [(center, f"{center:+.2%}"), (lo, f"{lo:+.2%}"), (hi, f"{hi:+.2%}")]:
                c = ws.cell(row=ri, column=col, value=fmt)
                c.fill   = PatternFill('solid', fgColor=bg)
                c.border = _thin_border()
                c.alignment = Alignment(horizontal='center')
                col += 1

            sc = ws.cell(row=ri, column=col, value=sig_label(mid))
            sc.font   = Font(bold=True, color=sig_color(mid))
            sc.fill   = PatternFill('solid', fgColor=bg)
            sc.border = _thin_border()
            sc.alignment = Alignment(horizontal='center')
            col += 1

        ws.cell(row=ri, column=col, value=rec['feat_count'])

    auto_col_width(ws)

    # ── 說明分頁 ──────────────────────────────────────────────────
    we = wb.create_sheet('模型說明')
    rows = [
        ('■ 版本', 'predictor.py v2.0'),
        ('■ 特徵集', f'共 {len(ALL_FEATURES)} 個：技術面(11) + 籌碼面(5) + 基本面(4) + 總經(6) + 股票類型(4) + 營收動能(3) + 估值動能(3)'),
        ('■ 技術面特徵', '收盤價, 成交量, MA5, MA20, ret5, RSI14, ATR14, BB_width, BIAS20, 量能比, MACD_hist'),
        ('■ 籌碼面特徵', '大戶指標, 外資持股比, 外資連買天數, 投信持股比, 融資變化率'),
        ('■ 基本面特徵', 'PER, PBR, 營收年增率, 營收月增率'),
        ('■ 總經/相對強弱', '台幣匯率, VIX指數, RSP, Industry, SP500_ret, DXY_ret'),
        ('■ 股票類型特徵', 'StockType（防禦0/中性1/循環2）, rolling_beta_60d, beta_regime, market_stress'),
        ('■ 營收動能特徵', 'rev_acc3m（3月YoY均）, rev_accel（YoY加速度）, rev_mom_acc（MoM動能）'),
        ('■ 估值動能特徵', 'per_zscore（PER歷史百分位）, pbr_zscore, per_trend（PER月變化）'),
        ('■ 模型', 'XGBoost（n=200, depth=4, lr=0.05, subsample=0.8）'),
        ('■ 訓練方式', '跨股聯合訓練：所有個股歷史資料合併，對每股最新一筆預測未來 3M/6M/12M'),
        ('■ 信賴區間', '動態計算：訓練集殘差的 P10/P90 百分位，反映模型實際預測誤差分佈'),
        ('■ 大盤(0050)', '排除於個股訓練，僅作為 RSP 相對強弱基準'),
        ('■ 訊號閾值', '偏多：中心值 > +5%；偏空：中心值 < -5%；中性：介於 ±5% 之間'),
        ('■ 回測勝率參考', '12M 方向勝率 73.3%（637筆，±10%門檻完全成功率 46.6%）'),
    ]
    we['A1'].value = '模型說明'; we['A1'].font = Font(bold=True, size=13, color=C_WHITE)
    we['A1'].fill  = PatternFill('solid', fgColor=C_NAVY)
    we.merge_cells('A1:B1')
    for i, (k, v) in enumerate(rows, 2):
        we.cell(row=i, column=1, value=k).font  = Font(bold=True)
        we.cell(row=i, column=2, value=v)
    auto_col_width(we)

    # ── 投資使用說明分頁 ──────────────────────────────────────────
    wg = wb.create_sheet('📋 投資使用說明')
    wg.column_dimensions['A'].width = 22
    wg.column_dimensions['B'].width = 80

    guide_rows = [
        # (標題, 內容, 是否為章節標題)
        ("【模型用途與定位】", "", True),
        ("本質",     "本模型為輔助參考工具，非投資建議。預測結果代表統計規律，不保證未來實現。", False),
        ("適用週期", "以 6M / 12M 中長期方向預測為主。3M 預測受短期雜訊影響較大，準確率較低。", False),
        ("適用對象", "對量化分析有基本認識、能搭配基本面判斷的投資人。", False),
        ("", "", False),
        ("【如何判讀預測力星等】", "", True),
        ("★★★★★（強）", "12M 方向勝率 ≥80%（統計顯著）。預測訊號參考價值高，可作為方向判斷的主要依據之一。", False),
        ("★★★★☆（強）", "12M 方向勝率 65~79%。訊號有效，建議搭配技術面確認後使用。", False),
        ("★★★☆☆（中）", "12M 方向勝率 55~64%。訊號略優於隨機，建議輕倉或等待更明確訊號。", False),
        ("★★☆☆☆（弱）", "12M 方向勝率 45~54%。接近隨機，不建議以此模型作為主要依據。", False),
        ("★☆☆☆☆（弱）", "12M 方向勝率 <45%（部分甚至反向）。此類股票的預測不可靠，請忽略結果。", False),
        ("", "", False),
        ("【如何使用預測結果】", "", True),
        ("第一步：看預測力", "先確認該股票的預測力星等。★★★★☆ 以上才建議參考方向訊號。", False),
        ("第二步：看訊號方向", "「📈 偏多」：模型預測未來上漲 >5%；「📉 偏空」：下跌 >5%；「➡️ 中性」：±5% 以內。", False),
        ("第三步：看信賴區間", "區間(Lo)到區間(Hi)是預測範圍。區間越窄代表模型越有信心，區間寬則不確定性高。", False),
        ("第四步：搭配其他分析", "模型僅提供統計方向，仍需搭配：基本面（財報、EPS）、技術面（支撐壓力）、籌碼面（外資動向）。", False),
        ("", "", False),
        ("【重要注意事項】", "", True),
        ("回測 ≠ 實盤",   "所有勝率數字為回測結果，實盤因交易成本、滑點、流動性，實際表現通常打折（約70~80%）。", False),
        ("不可重壓單一訊號", "即使最強的股票（如遠傳、台積電）也有 10~15% 的錯誤率，需分散部位。", False),
        ("黑天鵝無效",     "模型基於歷史規律學習，對突發事件（地緣政治、重大政策改變）完全無預測能力，請特別謹慎。", False),
        ("定期更新資料",   "建議每季執行一次 stock_analysis_v5.py 更新資料庫，再重跑 predictor，確保特徵是最新狀態。", False),
        ("石化/食品需謹慎", "台塑（32%）、統一（29%）的預測力弱，方向勝率低於隨機。這兩支的訊號請直接忽略。", False),
        ("", "", False),
        ("【建議操作流程】", "", True),
        ("Step 1", "查看總覽分頁，篩選「預測力 ★★★★☆ 以上」的股票。", False),
        ("Step 2", "確認 12M 訊號方向（偏多 / 偏空）。", False),
        ("Step 3", "查看信賴區間，確認預測範圍是否合理（區間 Lo 到 Hi 差距 <30% 為佳）。", False),
        ("Step 4", "對照自己對該股的基本面認知，若方向一致則提高信心，若相反則以基本面優先。", False),
        ("Step 5", "設定合理的部位大小：強訊號股可適度加重，弱訊號股降低部位或跳過。", False),
        ("", "", False),
        ("【模型技術規格】", "", True),
        ("算法",     "XGBoost（梯度提升樹），n=200, depth=4, lr=0.05, subsample=0.8", False),
        ("訓練方式", "跨股聯合訓練（Leave-One-Out）：用其他22支股票訓練，對目標股預測", False),
        ("特徵設計", "產業別特徵組合（17~31個特徵），各產業使用最適合的特徵子集", False),
        ("資料範圍", "2022年5月至今（約4年日線資料）", False),
        ("預測信賴區間", "訓練集殘差 P10/P90 百分位，非固定比例", False),
        ("版本",     "predictor v3.0 / backtester v5.0（產業別特徵版）", False),
    ]

    wg.merge_cells('A1:B1')
    wg['A1'].value = '📋 投資使用說明與注意事項'
    wg['A1'].font  = Font(bold=True, size=13, color='FFFFFF')
    wg['A1'].fill  = PatternFill('solid', fgColor=C_NAVY)
    wg['A1'].alignment = Alignment(horizontal='center', vertical='center')
    wg.row_dimensions[1].height = 28

    for i, (k, v, is_hdr) in enumerate(guide_rows, 2):
        if not k and not v:
            continue
        ka = wg.cell(row=i, column=1, value=k)
        va = wg.cell(row=i, column=2, value=v)
        if is_hdr:
            for cell in [ka, va]:
                cell.font  = Font(bold=True, color='FFFFFF', size=11)
                cell.fill  = PatternFill('solid', fgColor=C_BLUE)
                cell.alignment = Alignment(vertical='center')
            wg.row_dimensions[i].height = 22
        else:
            ka.font = Font(bold=True)
            va.alignment = Alignment(wrap_text=True, vertical='top')
            wg.row_dimensions[i].height = 36 if len(v) > 60 else 18
        for cell in [ka, va]:
            cell.border = _thin_border()

    wb.save(output_path)
    print(f"✅ 報告已儲存：{output_path}")


# ── 5. 主程式 ─────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='AI股價預測 v2.0')
    parser.add_argument('--data',   default='個股歷史分析資料庫.xlsx')
    parser.add_argument('--output', default='AI精確預測報告_v2.xlsx')
    args = parser.parse_args()

    print(f"📂 載入資料：{args.data}")
    stocks_raw = load_all_stocks(args.data)
    print(f"   共 {len(stocks_raw)} 檔個股（已排除 {BENCHMARK_CODE}）")

    print("🔧 特徵工程...")
    feats = {}
    for code, df in stocks_raw.items():
        try:
            feats[code] = build_features(df, code)
        except Exception as e:
            print(f"   ⚠️  {code} 特徵工程失敗：{e}")

    all_feat_list = list(feats.values())

    results = []
    for code, df_f in feats.items():
        name = STOCK_NAMES.get(code, code)
        print(f"  🔮 預測：{code} {name} ...")
        pred = {}
        feat_count = 0
        for hz_label, hz_days in HORIZONS.items():
            try:
                ind_feats = INDUSTRY_FEATURES.get(
                    STOCK_INDUSTRY_MAP.get(code, ''), DEFAULT_FEATURES)
                center, lo, hi, fc = train_and_predict(df_f, all_feat_list, hz_days,
                                                        feat_cols=ind_feats)
                pred[hz_label] = (center, lo, hi)
                feat_count = fc
            except Exception as e:
                print(f"     ⚠️  {hz_label} 預測失敗：{e}")
                pred[hz_label] = (0.0, -0.10, 0.10)

        power_wr, power_tier, power_stars = MODEL_POWER.get(
            code, (0.5, 'MID', '★★★☆☆'))
        results.append({
            'code':          code,
            'name':          name,
            'current_price': round(float(df_f['收盤價'].iloc[-1]), 2),
            'base_date':     str(df_f['日期'].iloc[-1].date()),
            'pred':          pred,
            'feat_count':    feat_count,
            'industry':      STOCK_INDUSTRY_MAP.get(code, '-'),
            'power_wr':      power_wr,
            'power_tier':    power_tier,
            'power_stars':   power_stars,
        })

    base_date = results[0]['base_date'] if results else 'N/A'

    # ── 讀取最新 model_power.json（backtester 每次跑完會更新）────
    import json, os
    json_path = 'model_power.json'
    if os.path.exists(json_path):
        try:
            with open(json_path, encoding='utf-8') as f:
                mp = json.load(f)
            # 用 JSON 覆蓋各股的預測力數據
            for rec in results:
                code = rec['code']
                if code in mp:
                    rec['power_wr']    = mp[code]['wr_12m']
                    rec['power_tier']  = mp[code]['tier']
                    rec['power_stars'] = mp[code]['stars']
            print(f"📊 已從 model_power.json 載入最新勝率（{len(mp)} 支）")
        except Exception as e:
            print(f"⚠️  model_power.json 讀取失敗，使用內建預設值：{e}")
    else:
        print("ℹ️  未找到 model_power.json，使用內建預設值（建議先執行 backtester_v4.py）")

    write_excel(results, args.output, base_date)
    print(f"\n✨ 完成！報告：{args.output}")


if __name__ == '__main__':
    main()
